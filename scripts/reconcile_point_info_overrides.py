"""One-off: push hand-typed Point Info room corrections back into the Master sheet.

When a tech types a room name straight into a Point Info 'LOCATION OF DEVICES' cell
(overwriting its `=Master!B{row}` formula) without updating the Master sheet, the
Master goes stale and the door chart (which reads only Master) misses the correction.

This script reconciles an existing workbook in place (Point Info wins): each literal
Point Info col-B override is written into the matching Master cell.

It edits the Master sheet's XML *surgically* inside the .xlsx zip, leaving every other
part byte-for-byte identical. A full openpyxl re-save is deliberately avoided — openpyxl
silently drops workbook parts it doesn't model (queryTables, connections, calcChain,
customXml, ...), which makes Excel flag the file as corrupt. The Point Info cells are
left as-is: they already show the corrected text, the door chart reads only Master, and
parse_dmp_worksheet now folds such literals in on import anyway.

A `*.bak` copy is written first. Usage:
    python scripts/reconcile_point_info_overrides.py "<path to .xlsx>"
"""
from __future__ import annotations

import re
import shutil
import sys
import zipfile
from pathlib import Path
from xml.sax.saxutils import escape

import openpyxl

_MASTER_A_REF_RE = re.compile(r"^=\s*Master!A(\d+)\s*$", re.IGNORECASE)


def _detect_overrides(path: Path) -> dict[int, tuple[str, str]]:
    """Master row -> (zone_label, override_text) for hand-typed Point Info col-B cells."""
    wb = openpyxl.load_workbook(str(path), data_only=False, read_only=True)
    out: dict[int, tuple[str, str]] = {}
    try:
        if "Master" not in wb.sheetnames:
            return out
        master = wb["Master"]
        master_label = {r: master.cell(r, 1).value for r in range(2, master.max_row + 1)}
        for name in wb.sheetnames:
            if "DMP 714-16 Point Info" not in name:
                continue
            ws = wb[name]
            for row in ws.iter_rows(min_row=4, max_row=19):
                a = row[0].value
                b = row[1].value if len(row) > 1 else None
                if not isinstance(a, str):
                    continue
                m = _MASTER_A_REF_RE.match(a.strip())
                if not m or not isinstance(b, str) or b.startswith("=") or not b.strip():
                    continue
                mrow = int(m.group(1))
                out[mrow] = (str(master_label.get(mrow, "")), b.strip())
    finally:
        wb.close()
    return out


def _resolve_master_sheet_xml(zin: zipfile.ZipFile) -> str:
    """Return the archive name of the Master sheet's worksheet XML."""
    wb_xml = zin.read("xl/workbook.xml").decode("utf-8")
    rels = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    relmap = dict(re.findall(r'Id="([^"]+)"[^>]*?Target="([^"]+)"', rels))
    for name, rid in re.findall(r'<sheet[^>]*?name="([^"]+)"[^>]*?r:id="([^"]+)"', wb_xml):
        if name == "Master":
            target = relmap[rid].lstrip("/")
            return target if target.startswith("xl/") else "xl/" + target
    raise ValueError("Master sheet not found in workbook.xml")


def _rewrite_cell_inline(xml: str, col: str, row: int, text: str) -> str:
    """Replace cell <col><row> with a self-contained inline string, preserving style."""
    ref = f"{col}{row}"
    pat = re.compile(r'<c r="%s"([^>]*?)(/>|>.*?</c>)' % re.escape(ref), re.S)

    def repl(m: re.Match) -> str:
        attrs = re.sub(r'\s+t="[^"]*"', "", m.group(1))  # drop any type attr
        body = escape(text)
        space = ' xml:space="preserve"' if text != text.strip() else ""
        return f'<c r="{ref}"{attrs} t="inlineStr"><is><t{space}>{body}</t></is></c>'

    new_xml, n = pat.subn(repl, xml, count=1)
    if n != 1:
        raise ValueError(f"cell {ref} not found exactly once in Master XML")
    return new_xml


def reconcile(path: str | Path) -> list[tuple[str, str, str]]:
    """Apply Point Info overrides to Master in `path`. Returns (zone, old, new) tuples."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(path)

    overrides = _detect_overrides(path)
    if not overrides:
        return []

    with zipfile.ZipFile(str(path)) as zin:
        master_name = _resolve_master_sheet_xml(zin)
        master_xml = zin.read(master_name).decode("utf-8")
        # snapshot old values for the report before mutating
        changes: list[tuple[str, str, str]] = []
        for mrow, (zone, text) in sorted(overrides.items()):
            old_m = re.search(r'<c r="B%d"[^>]*t="s"><v>(\d+)</v>' % mrow, master_xml)
            old = ""
            if old_m:
                ss = zin.read("xl/sharedStrings.xml").decode("utf-8")
                sis = re.findall(r"<si>(.*?)</si>", ss, re.S)
                idx = int(old_m.group(1))
                if idx < len(sis):
                    old = "".join(re.findall(r"<t[^>]*>(.*?)</t>", sis[idx], re.S))
            master_xml = _rewrite_cell_inline(master_xml, "B", mrow, text)
            changes.append((zone, old, text))

        infos = zin.infolist()
        payload = {info.filename: zin.read(info.filename) for info in infos}

    payload[master_name] = master_xml.encode("utf-8")

    backup = path.with_suffix(path.suffix + ".bak")
    shutil.copy2(path, backup)
    print(f"Backed up original -> {backup.name}")

    tmp = path.with_suffix(path.suffix + ".tmp")
    with zipfile.ZipFile(str(tmp), "w") as zout:
        for info in infos:  # preserve original order and per-entry compression
            zout.writestr(info, payload[info.filename])
    tmp.replace(path)
    return changes


def main(argv: list[str]) -> int:
    if len(argv) != 2:
        print(__doc__)
        return 2
    changes = reconcile(argv[1])
    if not changes:
        print("No hardcoded Point Info overrides found — nothing to do.")
        return 0
    print(f"Reconciled {len(changes)} zone(s) (Point Info -> Master):")
    for zone, old, new in changes:
        print(f"  {zone}: {old!r} -> {new!r}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
