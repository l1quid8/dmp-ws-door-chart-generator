"""
Generate a populated DMP Installation Worksheet from a design PDF.

Workflow: design PDF → OCR (if needed) → parse zone schedule + topology → build DMPDesign
→ prompt for metadata gaps → write DMP worksheet.

Usage:
    python scripts/generate_dmp_ws.py <design.pdf> [--searchable <path>] [--output <path>] [--non-interactive]
"""
from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter

# Make sibling modules importable
sys.path.insert(0, str(Path(__file__).parent))

from parse_zone_schedule import (parse_searchable_pdf, ZoneRecord, CombusLine,
                                 backfill_missing_expander_points)
from parse_dmp_worksheet import (
    DMPDesign, SiteInfo, RSP, Keypad, Splitter, PowerSupply, ZoneInfo
)
import re

from extract_topology import cluster_devices, extract_spans
from inject_door_chart import _slugify

# Phase 3 (vector-line edge detection) is optional — gracefully degrade if it's not in place.
try:
    from extract_topology import extract_full_topology  # type: ignore
    _PHASE3_AVAILABLE = True
except ImportError:
    _PHASE3_AVAILABLE = False


# Regex helpers for translating Phase-3 device IDs (e.g. "710-LX500-1") to DMP-format strings
_LX_BUS_RE = re.compile(r"710[\s\-]?LX(\d{3})[\s\-]?(\d+)", re.IGNORECASE)
_KP_NUM_RE = re.compile(r"710[\s\-]?KP[\s\-]?(\d+)", re.IGNORECASE)
_RSP_NUM_RE = re.compile(r"RSP\s*(\d+)", re.IGNORECASE)
_KEYPAD_NUM_RE = re.compile(r"KEYPAD\s*(\d+)", re.IGNORECASE)


def _norm_loc(s) -> str:
    """Normalize a location string for fuzzy matching ('Admin Bldg.  AV Room' → 'admin bldg av room')."""
    if not s:
        return ""
    return re.sub(r"[^\w]+", " ", str(s)).strip().lower()


def _apply_phase3_topology(design, edges, phase3_devices) -> int:
    """Populate splitter inputs/outputs from Phase-3 edges. Returns number of edges applied.

    DMP-worksheet output conventions:
      - Splitter outputs: 'RSP N', 'KEYPAD #N', 'To 710-LX500-X', 'To 710-KP-X', 'Spare'
      - LX splitter inputs: '500 BUS IN FROM XR/550' (top of bus) or 'From 710-LX500-X' (downstream)
      - KP splitter inputs: 'KEYPAD BUS IN FROM XR/550' or 'From 710-KP-X'
    """
    if not edges:
        return 0

    # Build location → DMP splitter id maps (separately for LX vs KP)
    loc_to_lx_id: dict[str, str] = {}
    loc_to_kp_id: dict[str, str] = {}
    for s in design.splitters:
        key = _norm_loc(s.location)
        if not key:
            continue
        if s.splitter_type == "LX":
            loc_to_lx_id.setdefault(key, s.id)
        else:
            loc_to_kp_id.setdefault(key, s.id)

    # If multiple LX splitters share a location (common: 3 splitters in one A/V room), the
    # location-only map collapses them. Fall back to ordering: index Phase-3 LX devices by
    # their parsed bus+slot, then pair them with DMP LX splitters in order.
    phase3_lx = sorted(
        [d for d in phase3_devices if d.kind == "SPLITTER" and _LX_BUS_RE.search(d.id or "")],
        key=lambda d: (
            int(_LX_BUS_RE.search(d.id).group(1)),
            int(_LX_BUS_RE.search(d.id).group(2)),
        ),
    )
    phase3_kp = sorted(
        [d for d in phase3_devices if d.kind == "SPLITTER" and _KP_NUM_RE.search(d.id or "")],
        key=lambda d: int(_KP_NUM_RE.search(d.id).group(1)),
    )
    dmp_lx = [s for s in design.splitters if s.splitter_type == "LX"]
    dmp_kp = [s for s in design.splitters if s.splitter_type == "KP"]

    phase3_id_to_dmp_id: dict[str, str] = {}
    # Direct positional pairing (works when Phase 3 found the same number of splitters as we created)
    for p3, dmp in zip(phase3_lx, dmp_lx):
        phase3_id_to_dmp_id[p3.id] = dmp.id
    for p3, dmp in zip(phase3_kp, dmp_kp):
        phase3_id_to_dmp_id[p3.id] = dmp.id

    # Bus number for top-of-bus inputs — keyed by phase3 device id
    phase3_id_to_bus: dict[str, int] = {}
    for d in phase3_lx:
        m = _LX_BUS_RE.search(d.id or "")
        if m:
            phase3_id_to_bus[d.id] = int(m.group(1))

    def _output_desc(dst_dev) -> Optional[str]:
        """Format a Phase-3 destination device as a DMP-style output string."""
        if dst_dev.kind == "RSP":
            m = _RSP_NUM_RE.search(dst_dev.id or "")
            return f"RSP {m.group(1)}" if m else None
        if dst_dev.kind == "KEYPAD":
            m = _KEYPAD_NUM_RE.search(dst_dev.id or "")
            return f"KEYPAD #{m.group(1)}" if m else None
        if dst_dev.kind == "SPLITTER":
            dmp_id = phase3_id_to_dmp_id.get(dst_dev.id)
            return f"To {dmp_id}" if dmp_id else None
        return None

    # Bucket edges by source DMP splitter id
    outputs_by_src: dict[str, list[str]] = {}
    inputs_by_dst: dict[str, str] = {}  # dst_dmp_id -> input description (splitter→splitter only)
    n_applied = 0

    for edge in edges:
        src_dmp_id = phase3_id_to_dmp_id.get(edge.src.id)
        dst_dmp_id = phase3_id_to_dmp_id.get(edge.dst.id)

        # Splitter → anything: this is an output of src
        if src_dmp_id:
            desc = _output_desc(edge.dst)
            if desc:
                outputs_by_src.setdefault(src_dmp_id, []).append(desc)
                n_applied += 1

        # Splitter → splitter: also record the input for dst — 'From {full diagram id}'
        if src_dmp_id and dst_dmp_id and edge.src.kind == "SPLITTER":
            inputs_by_dst[dst_dmp_id] = f"From {src_dmp_id}"

    # Apply to DMP splitters
    for splitter in design.splitters:
        outs = list(dict.fromkeys(outputs_by_src.get(splitter.id, [])))  # dedupe, keep order
        # A splitter has only 3 outputs. If extraction over-subscribed it, keep
        # the structural outputs (chain links to other splitters, RSP feeds) and
        # let excess keypads drop — keypad distribution is the least certain part
        # of extraction and is corrected in the review step.
        if len(outs) > 3:
            outs.sort(key=lambda o: 0 if str(o).startswith(("To ", "RSP ")) else 1)
        # Pad to 3 outputs with "Spare" (DMP convention)
        while len(outs) < 3:
            outs.append("Spare")
        splitter.outputs = outs[:3]

        # Inputs: explicit "From X" if splitter→splitter, otherwise "{bus} BUS IN FROM XR/550"
        if splitter.id in inputs_by_dst:
            input_key = "LX-Bus In" if splitter.splitter_type == "LX" else "KP-Bus In"
            splitter.inputs = {input_key: inputs_by_dst[splitter.id]}
        else:
            if splitter.splitter_type == "LX":
                # Find the corresponding Phase 3 device to recover the bus number
                bus = None
                for p3, dmp in zip(phase3_lx, dmp_lx):
                    if dmp.id == splitter.id:
                        bus = phase3_id_to_bus.get(p3.id)
                        break
                if bus is None:
                    bus = 500  # safe default (most common single-bus deployment)
                splitter.inputs = {"LX-Bus In": f"{bus} BUS IN FROM XR/550"}
            else:
                splitter.inputs = {"KP-Bus In": "KEYPAD BUS IN FROM XR/550"}

    return n_applied


def _phase3_topology_complete(design: DMPDesign) -> bool:
    """True if riser-extracted splitter I/O covers every RSP and every
    non-service keypad as a splitter output — i.e. it can be trusted over the
    auto-derive convention. If anything is missing, the convention is used."""
    produced: set[str] = set()
    for s in design.splitters:
        for o in (s.outputs or []):
            produced.add(str(o).strip())
    for rsp in design.rsps:
        if f"RSP {rsp.number}" not in produced:
            return False
    for kp in design.keypads:
        if kp.number != 1 and f"KEYPAD #{kp.number}" not in produced:
            return False
    return True


_SHEET_RE = re.compile(r"INT[-\s]?(\d+\.\d+)", re.IGNORECASE)
RISER_SHEET = "5.0"  # INT-5.0 is the riser diagram in the C1 standard sheet set


def _page_sheet_number(text: str) -> str | None:
    """The page's own sheet number, i.e. the sole INT-x.x ref in its title block.

    Returns None for the cover/index page (which lists every sheet, so many refs)
    or unlabeled pages, so those never match a specific sheet.
    """
    nums = {m.group(1) for m in _SHEET_RE.finditer(text)}
    return next(iter(nums)) if len(nums) == 1 else None


def _detect_riser_page(pdf_path: Path) -> int:
    """Find the page index of the riser-diagram sheet (INT-5.0).

    Primary: pick the page whose own title block is INT-5.0 — deterministic and
    independent of page content. (A naive search for "INT-5.0" is not enough: the
    cover/index sheet also lists it, so we match only the page where it's the SOLE
    sheet number.)

    Fallback: SPLITTER anchors (`710-LX500-N`, `710-KP-N`) cluster on the riser, so
    the page with the most splitter anchors (tiebroken by total device count) is the
    best guess. Used only when no page is labeled INT-5.0 (renumbered set, image-only
    title block, OCR gap). This is the historical heuristic, and it is fragile — e.g.
    a dense siteplan can tie the riser on splitter anchors — which is exactly why the
    sheet-number path is preferred.
    """
    from extract_topology import extract_spans, classify_label
    import fitz

    best_page = 0
    best_splitters = -1
    best_total = -1
    doc = fitz.open(str(pdf_path))
    try:
        for i in range(len(doc)):
            try:
                spans = extract_spans(pdf_path, page_idx=i)
            except Exception:
                continue
            # Primary path: this page's title block is the riser sheet.
            page_text = " ".join(sp.text for sp in spans)
            if _page_sheet_number(page_text) == RISER_SHEET:
                return i
            splitter_count = 0
            total_count = 0
            for sp in spans:
                cls = classify_label(sp.text)
                if cls.startswith("device:"):
                    total_count += 1
                    if cls == "device:SPLITTER":
                        splitter_count += 1
            # Prefer pages with splitter anchors; tiebreak on total device count.
            if (splitter_count, total_count) > (best_splitters, best_total):
                best_splitters = splitter_count
                best_total = total_count
                best_page = i
    finally:
        doc.close()
    print(
        f"_detect_riser_page: no page labeled INT-{RISER_SHEET}; "
        f"falling back to splitter-anchor heuristic (page {best_page}).",
        file=sys.stderr,
    )
    return best_page


def _auto_derive_splitter_io(design) -> None:
    """Populate splitter inputs/outputs from RSP/Keypad lists when Phase 3 yields no edges.

    LX chain convention (using IA-diagram IDs):
      710-LX500-1: in='500 BUS IN FROM XR/550'; outputs=['RSP 1', 'To 710-LX500-N', ...] padded with 'Spare'
      710-LX500-N (N>=2): in='From 710-LX500-1'; outputs = remaining RSPs round-robin, padded
    KP convention (chained, mirrors LX):
      710-KP-1: in='KEYPAD BUS IN FROM XR/550'; outputs=keypads + 'To 710-KP-N' links, padded
      710-KP-N (N>=2): in='From 710-KP-1'; outputs = remaining keypads round-robin, padded
    """
    lx = [s for s in design.splitters if s.splitter_type == "LX"]
    kp = [s for s in design.splitters if s.splitter_type == "KP"]
    rsps = sorted(design.rsps, key=lambda r: r.number)

    if lx:
        lx[0].inputs = {"LX-Bus In": "500 BUS IN FROM XR/550"}
        outs = [f"RSP {rsps[0].number}"] if rsps else []
        for chained in lx[1:3]:
            outs.append(f"To {chained.id}")
        while len(outs) < 3:
            outs.append("Spare")
        lx[0].outputs = outs[:3]

        downstream = lx[1:]
        buckets: dict[str, list[str]] = {s.id: [] for s in downstream}
        for i, r in enumerate(rsps[1:]):
            if not downstream:
                break
            buckets[downstream[i % len(downstream)].id].append(f"RSP {r.number}")
        upstream_id = lx[0].id
        for s in downstream:
            s.inputs = {"LX-Bus In": f"From {upstream_id}"}
            outs = buckets[s.id]
            while len(outs) < 3:
                outs.append("Spare")
            s.outputs = outs[:3]

        # Warn if RSPs would be dropped (more RSPs than splitter outputs can hold)
        capacity = 1 + 3 * len(downstream)  # LX-1 holds RSP-1; downstream holds 3 each
        if len(rsps) > capacity:
            print(f"  WARNING: {len(rsps)} RSPs but splitter capacity is {capacity} — "
                  f"{len(rsps) - capacity} RSP(s) not assigned to a splitter output.")

    if kp:
        # KP splitters daisy-chain off kp[0] exactly like the LX bus above —
        # kp[0] is fed from the XR/550, every downstream KP splitter is fed
        # "From 710-KP-1". (Earlier code parallelized them off the MSP, which
        # produced wrong wiring on real designs — KP buses chain like LX buses.)
        non_service = [k for k in design.keypads if k.number != 1]
        downstream = kp[1:]

        # kp[0]: keypads fill the slots not needed for chain links, then one
        # "To 710-KP-N" link per downstream splitter, padded with "Spare".
        kp[0].inputs = {"KP-Bus In": "KEYPAD BUS IN FROM XR/550"}
        kp0_keypad_slots = max(0, 3 - len(downstream))
        outs = [f"KEYPAD #{k.number}" for k in non_service[:kp0_keypad_slots]]
        for chained in downstream:
            outs.append(f"To {chained.id}")
        while len(outs) < 3:
            outs.append("Spare")
        kp[0].outputs = outs[:3]

        # Downstream KP splitters chain from kp[0] and carry the remaining keypads.
        kp_buckets: dict[str, list[str]] = {s.id: [] for s in downstream}
        for i, k in enumerate(non_service[kp0_keypad_slots:]):
            if not downstream:
                break
            kp_buckets[downstream[i % len(downstream)].id].append(f"KEYPAD #{k.number}")
        for s in downstream:
            s.inputs = {"KP-Bus In": f"From {kp[0].id}"}
            outs = kp_buckets[s.id]
            while len(outs) < 3:
                outs.append("Spare")
            s.outputs = outs[:3]

        # Warn if keypads exceed the chained capacity (kp[0] slots + 3 per downstream).
        capacity = kp0_keypad_slots + 3 * len(downstream)
        if len(non_service) > capacity:
            print(f"  WARNING: {len(non_service)} non-service keypads but KP-bus capacity "
                  f"is {capacity} — {len(non_service) - capacity} keypad(s) not assigned to a "
                  f"splitter output.")


from paths import resource_path, output_dir

DEFAULT_TEMPLATE = resource_path("DMP Installation Worksheet_template_blank.xlsx")
DEFAULT_OUTPUT_DIR = output_dir()


def resolve_original_pdf(pdf_path: Path) -> Path:
    """If pdf_path is `<name>_searchable.pdf`, return `<name>.pdf` if it exists.

    The riser-diagram page has vector text that PyMuPDF reads accurately, but
    OCR can mangle splitter labels (e.g. Academy's '710-LX500-1' / '710-LX500-2'
    don't survive ocrmypdf round-trip and become invisible to PyMuPDF on the
    searchable PDF). Topology extraction must run against the original PDF.
    """
    name = pdf_path.name
    suffix = "_searchable.pdf"
    if name.lower().endswith(suffix):
        original = pdf_path.with_name(name[: -len(suffix)] + ".pdf")
        if original.exists():
            return original
    return pdf_path


def ensure_searchable_pdf(pdf_path: Path, searchable_override: Optional[Path] = None) -> Path:
    """Return a searchable PDF, using OCR if needed."""
    if searchable_override:
        if searchable_override.exists():
            return searchable_override
        raise FileNotFoundError(f"Searchable PDF not found: {searchable_override}")

    # Check if input is already searchable
    import fitz
    try:
        doc = fitz.open(str(pdf_path))
        for p in doc:
            t = p.get_text("text")
            if "MOTION DETECTOR ZONE SCHEDULE" in t.upper() and len(t) > 2000:
                doc.close()
                return pdf_path
        doc.close()
    except Exception:
        pass

    # Need to OCR
    try:
        from prepare_pdf import prepare
        print(f"  Converting {pdf_path.name} to searchable (ocrmypdf)...")
        return prepare(pdf_path)
    except ImportError:
        raise RuntimeError(
            "PDF is not searchable and ocrmypdf/prepare_pdf not available. "
            "Provide --searchable <path> to an already-OCR'd PDF."
        )


# --- Source-data conflict detection -----------------------------------------
#
# An RSP's room is printed in two independent places on the CAD design: the
# COMBUS LINES table (the service-panel row) and the zone schedule (the RSP's
# A/C-loss and battery-trouble supervisory zones, which monitor that RSP's own
# power supply and are therefore in the same room as the RSP). When the CAD team
# mistypes one — e.g. RSP3 listed as "CLASSROOM 17" in COMBUS LINES but its
# supervisory zones say "ROOM 18" — the generator should not silently pick one.
# It collects the mismatch and escalates it to the user (CLI prompt / GUI dialog).

@dataclass
class LocationConflict:
    """A device location parsed inconsistently from two independent PDF sources."""
    kind: str           # "RSP" | "KEYPAD"
    number: int
    label: str          # human-readable, e.g. "RSP 3 location"
    options: list       # list[tuple[value: str, source: str]]


_NUM_TOKEN_RE = re.compile(r"\d+")
_FLOOR_NUM_RE = re.compile(r"\d+\s*(?:st|nd|rd|th)\b|\d+\s*(?:st|nd|rd|th)?\s*(?:flr|floor)\b", re.I)


def _significant_numbers(loc: str) -> set:
    """Numeric tokens in a location string, excluding floor ordinals
    ('1ST FLR', '2nd floor', ...). A room-number typo shows up here."""
    if not loc:
        return set()
    return set(_NUM_TOKEN_RE.findall(_FLOOR_NUM_RE.sub(" ", str(loc))))


def detect_location_conflicts(design: DMPDesign, parsed) -> list:
    """Cross-check each RSP's service-panel location (COMBUS LINES table) against
    the location of its supervisory zones (the A/C-loss / battery-trouble zones,
    which are co-located with the RSP). Flags a conflict when the room NUMBER
    disagrees — the CAD-typo class a technician has to fix by hand. Identical
    locations and wording-only differences are not escalated. `parsed` is the
    ParsedSchedule from parse_searchable_pdf. Returns a list of LocationConflict.
    """
    # Full supervisory-zone location per RSP number, from the zone schedule.
    sup_by_rsp: dict[int, str] = {}
    for z in getattr(parsed, "zones", None) or []:
        if not (z.is_ps_ac or z.is_ps_batt) or not z.room:
            continue
        loc = " ".join(p for p in (z.building, z.floor, z.room) if p).strip()
        if loc:
            sup_by_rsp.setdefault(z.rsp, loc)

    conflicts: list = []
    for rsp in design.rsps:
        combus_loc = rsp.location
        sup_loc = sup_by_rsp.get(rsp.number)
        if not combus_loc or not sup_loc:
            continue
        if _norm_loc(combus_loc) == _norm_loc(sup_loc):
            continue  # agree
        combus_nums = _significant_numbers(combus_loc)
        sup_nums = _significant_numbers(sup_loc)
        if not combus_nums or not sup_nums or combus_nums == sup_nums:
            continue  # need a real room-number disagreement (not missing/wording)
        conflicts.append(LocationConflict(
            kind="RSP", number=rsp.number,
            label=f"RSP {rsp.number} location",
            options=[(combus_loc, "COMBUS LINES table"),
                     (sup_loc, "supervisory zones (A/C-loss & battery)")],
        ))

    return conflicts


def apply_location_conflict(design: DMPDesign, conflict: LocationConflict, value: str) -> None:
    """Write the user-chosen location onto every design object that shares the
    conflicted device number (RSP also updates its matching power supply)."""
    if conflict.kind == "RSP":
        for rsp in design.rsps:
            if rsp.number == conflict.number:
                rsp.location = value
        for ps in design.power_supplies:
            if ps.number == conflict.number:
                ps.location = value
    elif conflict.kind == "KEYPAD":
        for kp in design.keypads:
            if kp.number == conflict.number:
                kp.location = value


def _expander_model_for_count(zone_count: int) -> str:
    """Infer the expander model from how many zones the PDF lists for the module.

    The zone schedule lists every physical point per module (motion + spares + the
    two paired-power-supply supervisory points), so the count is the port count:
    a 714-8 carries 8, a 714-16 carries 16. Anything 8 or fewer is a 714-8.
    """
    return "714-8" if zone_count <= 8 else "714-16"


def _ps_relays(zone_nums: list[int], model: str, rsp_num: int) -> dict[int, str]:
    """Relay assignments for an expander's paired 505-12G power supply.

    The module's last two physical zones supervise the PS (A/C-loss then battery),
    so the trouble-zone labels must reference the actual top two zone numbers — not
    a fixed +14/+15 offset, which only ever held for a full 16-zone module.
    """
    if not zone_nums:
        return {}
    sv = sorted(zone_nums)
    ac, batt = (sv[-2], sv[-1]) if len(sv) >= 2 else ("??", "??")
    return {
        1: "12v DC Output to Terminal Strip",
        2: f"AC Trouble Zone {ac} ({model} Expander #{rsp_num})",
        3: f"Battery Trouble Zone {batt} ({model} Expander #{rsp_num})",
        4: "Battery 12V",
    }


def build_dmp_design_from_pdf(
    searchable_pdf: Path,
    design_pdf: Path,
    non_interactive: bool = False,
    prompt_routing: bool = False,
) -> DMPDesign:
    """Parse PDF data and build a DMPDesign object with user-prompted gaps."""
    # Parse zone schedule from searchable PDF
    parsed = parse_searchable_pdf(searchable_pdf)
    school_name = parsed.school_info.get("school_name", "Unknown School")
    school_code = parsed.school_info.get("loc_code", "")

    print(f"  School: {school_name} (code: {school_code})")
    print(f"  RSPs: {sum(1 for c in parsed.combus_lines if c.kind == 'RSP')}")
    print(f"  Keypads: {sum(1 for c in parsed.combus_lines if c.kind == 'KEYPAD')}")
    print(f"  Zones: {len(parsed.zones)}")

    # Extract topology (devices + locations) from the riser-diagram page. Auto-detect
    # the page by finding the one with the most splitter/RSP/keypad device anchors —
    # the page index varies per design (INT-5.0 may be page 6, 7, 10, etc.).
    devices = []
    riser_edges: list = []
    splitter_on_rsp: dict = {}
    riser_page = _detect_riser_page(design_pdf)
    try:
        from extract_topology import merge_multiline_locations, merge_horizontal_locations
        spans = extract_spans(design_pdf, page_idx=riser_page)
        spans = merge_multiline_locations(spans)
        spans = merge_horizontal_locations(spans)
        devices = cluster_devices(spans)
        print(f"  Devices from topology (page {riser_page}): {len(devices)}")
    except Exception as e:
        print(f"  Warning: could not extract topology from page {riser_page}: {e}")

    # Reconstruct the splitter/keypad wiring from the riser-diagram vector lines.
    if devices:
        try:
            from extract_topology import (extract_line_segments,
                compute_device_footprints, reconstruct_edges)
            segments = extract_line_segments(design_pdf, riser_page)
            footprints, splitter_on_rsp = compute_device_footprints(
                devices, segments, design_pdf, riser_page)
            riser_edges = reconstruct_edges(segments, devices, spans, footprints=footprints)
            print(f"  Riser edges extracted: {len(riser_edges)}")
        except Exception as e:
            print(f"  Warning: riser edge extraction failed: {e}")

    # Build DMPDesign structure
    design = DMPDesign()

    # Site info
    design.site_info.school_name = school_name
    design.site_info.school_code = school_code
    design.site_info.address_line1 = parsed.school_info.get("address_line1")
    design.site_info.address_line2 = parsed.school_info.get("address_line2")

    # The school phone isn't printed on the design PDF, so look it up live from
    # the LAUSD directory (school_lookup) by street number + ZIP. Fill only when
    # blank so a re-imported/edited value is preserved; editable in the SITE tab.
    if not (design.site_info.phone or "").strip():
        from school_lookup import lookup_phone
        phone = lookup_phone(
            design.site_info.school_name or "",
            design.site_info.address_line1 or "",
            design.site_info.address_line2 or "",
        )
        if phone:
            design.site_info.phone = phone
    msp_location = None
    for dev in devices:
        if dev.kind == "MSP" and dev.location:
            msp_location = dev.location
            break
    # If we didn't find MSP in topology, try to use first RSP location or MSP from first combus line
    if not msp_location:
        for combus_line in parsed.combus_lines:
            if combus_line.kind == "RSP" and combus_line.n == 1:
                msp_location = f"{combus_line.building} {combus_line.floor} {combus_line.room}"
                break
    design.site_info.xr550_location = msp_location or "TBD"

    # RSPs + zones assignment
    # Build a map of RSP# -> zone count from combus_lines and zone schedule
    # ONLY create RSPs for those explicitly listed in COMBUS LINES
    # Collect RSP numbers from COMBUS LINES
    rsp_combus_map: dict[int, CombusLine] = {}
    for combus_line in parsed.combus_lines:
        if combus_line.kind == "RSP":
            rsp_combus_map[combus_line.n] = combus_line

    # OCR routinely drops SPARE rows from the large-format zone schedule, leaving
    # holes in an expander's contiguous point block (e.g. Shirley RSP1 lost the
    # Z508-Z514 spares). Rebuild them deterministically from each installed
    # module's point count so every physical point lands on the worksheet/door
    # chart. Scoped to installed RSPs so a stray zone can't spawn a phantom block.
    parsed.zones, backfilled = backfill_missing_expander_points(
        parsed.zones, installed_rsps=set(rsp_combus_map.keys()))
    if backfilled:
        print(f"  Backfilled {backfilled} dropped spare point(s) across "
              f"{len(rsp_combus_map)} expander(s)")

    rsp_zones: dict[int, list[ZoneRecord]] = {}
    for zone in parsed.zones:
        if zone.rsp not in rsp_zones:
            rsp_zones[zone.rsp] = []
        rsp_zones[zone.rsp].append(zone)

    for rsp_num in sorted(rsp_combus_map.keys()):
        combus_line = rsp_combus_map[rsp_num]
        zones = rsp_zones.get(rsp_num, [])
        # Include ALL zones assigned to this RSP (including spares and PS supervisory)
        zone_nums = [int(z.zone[1:]) for z in zones]
        model = _expander_model_for_count(len(zone_nums))

        design.rsps.append(RSP(
            number=rsp_num,
            location=f"{combus_line.building} {combus_line.floor} {combus_line.room}",
            zones=zone_nums,
            model=model,
        ))
        design.power_supplies.append(PowerSupply(
            number=rsp_num,
            location=f"{combus_line.building} {combus_line.floor} {combus_line.room}",
            relays=_ps_relays(zone_nums, model, rsp_num),
        ))

    # Keypads: KP#1 = MSP service KP (from first RSP location), rest from combus_lines
    # Add KP#1 service keypad first (from MSP/XR550 location)
    design.keypads.append(Keypad(
        number=1,
        source="MSP",
        location=f"{design.site_info.xr550_location} (Service Keypad)",
        global_keypad=True,
    ))

    # Add keypads from COMBUS LINES (excluding the logical KP#1)
    # For now, assume all non-service keypads are fed by KP-710-1 (can be modified interactively)
    for combus_line in parsed.combus_lines:
        if combus_line.kind != "KEYPAD":
            continue
        kp_num = combus_line.n
        # Skip if this is being used as KP#1
        if kp_num == 1:
            continue
        design.keypads.append(Keypad(
            number=kp_num,
            source="710-KP-1",  # default: fed by first KP splitter (parse_dmp_worksheet will normalize format)
            location=f"{combus_line.building} {combus_line.floor} {combus_line.room}",
            global_keypad=True,
        ))

    # Splitters: from topology devices marked "SPLITTER"
    # Collect by type and renumber consistently
    splitters_by_type: dict[str, list] = {"LX": [], "KP": []}
    for dev in devices:
        if dev.kind == "SPLITTER":
            splitter_type = "KP" if "KP" in dev.id.upper() else "LX"
            splitters_by_type[splitter_type].append(dev)

    # Sort by location and assign IDs
    lx_devs = sorted(splitters_by_type["LX"], key=lambda d: d.location or "")
    kp_devs = sorted(splitters_by_type["KP"], key=lambda d: d.location or "")

    # Preserve the IA-diagram IDs (e.g. '710-LX500-1', '710-KP-1') extracted from the
    # riser by extract_topology, so the splitter sheet IDs match the diagram exactly.
    # Fall back to a synthesized ID only if the device has no id (shouldn't happen).
    for i, dev in enumerate(lx_devs, 1):
        design.splitters.append(Splitter(
            id=dev.id or f"710-LX500-{i}",
            splitter_type="LX",
            location=dev.location or "TBD",
            inputs={},
            outputs=[],
        ))

    for i, dev in enumerate(kp_devs, 1):
        design.splitters.append(Splitter(
            id=dev.id or f"710-KP-{i}",
            splitter_type="KP",
            location=dev.location or "TBD",
            inputs={},
            outputs=[],
        ))

    # Splitter locations: a splitter drawn on an RSP box is in that RSP's room.
    # This corrects the often-truncated location pulled from the riser label.
    if splitter_on_rsp:
        rsp_loc = {r.number: r.location for r in design.rsps}
        for s in design.splitters:
            rsp_id = splitter_on_rsp.get(s.id)
            if rsp_id:
                m = re.search(r"\d+", rsp_id)
                if m and int(m.group()) in rsp_loc and rsp_loc[int(m.group())]:
                    s.location = rsp_loc[int(m.group())]

    # Splitter I/O: prefer the wiring reconstructed from the riser diagram; fall
    # back to the deterministic auto-derive convention when riser extraction
    # didn't cover the whole design.
    if design.splitters:
        applied = _apply_phase3_topology(design, riser_edges, devices) if riser_edges else 0
        if applied and _phase3_topology_complete(design):
            design.topology_source = "riser"
            print(f"  Splitter I/O from riser extraction ({applied} edges applied)")
        else:
            _auto_derive_splitter_io(design)
            design.topology_source = "auto-derived"
            print(f"  Auto-derived splitter I/O for {len(design.splitters)} splitters")

    # Zones: create ZoneInfo for each zone in the schedule
    for zone in parsed.zones:
        z_num = int(zone.zone[1:])
        design.zones.append(ZoneInfo(
            number=z_num,
            location=zone.room or "",
            device_type="Motion" if not (zone.is_spare or zone.is_ps_ac or zone.is_ps_batt) else (
                "Supervisory" if (zone.is_ps_ac or zone.is_ps_batt) else "Spare"
            ),
            partition=1,
        ))

    # Cross-check locations parsed from independent PDF sources and surface any
    # mismatches. In a GUI run (non_interactive) the conflicts stay on the design
    # so the app can show a resolution dialog before generation.
    design.conflicts = detect_location_conflicts(design, parsed)
    if design.conflicts and non_interactive:
        for c in design.conflicts:
            print("  CONFLICT (needs review): " + c.label + " — " +
                  " vs ".join(f"{v!r} [{src}]" for v, src in c.options))

    # Interactive prompts for gaps (unless --non-interactive)
    if not non_interactive:
        if design.conflicts:
            print("\n=== Source-data inconsistencies (the CAD prints disagree) ===")
            for c in list(design.conflicts):
                print(f"\n{c.label}:")
                for i, (value, source) in enumerate(c.options, 1):
                    print(f"  {i}. {value}   [{source}]")
                choice = input(
                    f"  Choose 1-{len(c.options)}, or type the correct value: "
                ).strip()
                resolved: Optional[str] = None
                if choice.isdigit() and 1 <= int(choice) <= len(c.options):
                    resolved = c.options[int(choice) - 1][0]
                elif choice:
                    resolved = choice
                if resolved is not None:
                    apply_location_conflict(design, c, resolved)
                    print(f"  -> using: {resolved}")
            design.conflicts = []

        print("\n=== Metadata Gaps (press Enter to skip) ===")

        # School code — show auto-extracted value as the default
        current_code = design.site_info.school_code or ""
        prompt = f"School code [{current_code}]? " if current_code else "School code? "
        code = input(prompt).strip()
        if code:
            design.site_info.school_code = code

        phone = input("Main phone number? ").strip()
        if phone:
            design.site_info.phone = phone
        tech = input("Install tech name? ").strip()
        if tech:
            design.site_info.install_tech = tech

        install_date = input("Install date? ").strip()
        if install_date:
            design.site_info.install_date = install_date

        ip = input("IP address? ").strip()
        if ip:
            design.site_info.ip_address = ip

        gateway = input("Default gateway? ").strip()
        if gateway:
            design.site_info.default_gateway = gateway

        # Splitter I/O routing — opt-in via --prompt-routing. Shows the derived
        # wiring (from the riser, or the convention) and lets the user correct
        # each output; press Enter to keep the derived value.
        if prompt_routing and design.splitters:
            print(f"\n=== Splitter I/O Routing — derived from: "
                  f"{design.topology_source or 'auto-derived'} ===")
            for splitter in design.splitters:
                inp = list(splitter.inputs.values())[0] if splitter.inputs else ""
                print(f"\n{splitter.id} (at {splitter.location})  input: {inp}")
                rsp_names = [f"RSP {r.number}" for r in design.rsps]
                kp_names = [f"KEYPAD #{k.number}" for k in design.keypads if k.number != 1]
                splitter_names = [f"To {s.id}" for s in design.splitters if s.id != splitter.id]
                choices = rsp_names + kp_names + splitter_names + ["Spare"]
                print(f"  choices: {', '.join(choices)}")
                outs = list(splitter.outputs or [])
                for i in range(3):
                    cur = outs[i] if i < len(outs) else "Spare"
                    choice = input(f"  Output {i + 1} [{cur}]? ").strip()
                    if i < len(outs):
                        outs[i] = choice if choice else cur
                    elif choice:
                        outs.append(choice)
                splitter.outputs = (outs + ["Spare", "Spare", "Spare"])[:3]

    return design


def _write_cell_safe(ws, cell_ref: str, value) -> None:
    """Write to a cell, handling merged cells gracefully."""
    if value is None:
        return
    try:
        ws[cell_ref] = value
    except (AttributeError, TypeError):
        # Cell may be merged; skip
        pass


def dmp_filename(school_slug: str, stamp: str | None = None,
                 date_str: str | None = None) -> str:
    """Output filename for a generated DMP worksheet.

    DRAFT/FINAL goes in the name so completion status is unambiguous when
    techs browse the output folder.
    """
    from datetime import date as _date
    d = date_str or _date.today().isoformat()
    tag = f"_{stamp}" if stamp else ""
    return f"{school_slug}_dmp{tag}_{d}.xlsx"


def write_dmp_xlsx(design: DMPDesign, template_path: Path, output_path: Path,
                   stamp: str | None = None) -> None:
    """Copy the blank template and populate all sheets with design data.

    stamp: "DRAFT" adds a visible not-for-install banner and a machine-readable
    DMPStatus doc property (drafts are refused on re-import — the session is
    the source of truth); "FINAL" adds only the doc property.
    """
    import shutil
    shutil.copy(template_path, output_path)

    wb = openpyxl.load_workbook(output_path)

    # Normalize each expander's model + paired-PS relay text from its LIVE zone count.
    # Zone count == port count in this data model (add_expander materializes exactly
    # EXPANDER_MODELS[model] zones; the PDF lists every port). A worksheet re-imported from an
    # older/pre-fix file may carry a stale "714-16" in the Exp Mod text (and wrong supervisory
    # zones in the relay labels) even for an 8-port module; deriving from the zone count keeps
    # the Exp Mod sheet, LX-bus row, Point Info source labels, and PS relays consistent
    # regardless of how the design was loaded (PDF, xlsx re-import, or editor).
    for rsp in design.rsps:
        if rsp.zones:
            rsp.model = _expander_model_for_count(len(rsp.zones))
    _rsp_by_num = {r.number: r for r in design.rsps}
    for ps in design.power_supplies:
        rsp = _rsp_by_num.get(ps.number)
        if rsp and rsp.zones:
            ps.relays = _ps_relays(rsp.zones, rsp.model, ps.number)

    # SITE INFO sheet
    ws = wb["SITE INFO"]
    if stamp == "DRAFT":
        from datetime import date as _date
        _write_cell_safe(
            ws, "A1",
            f"*** DRAFT — NOT FOR INSTALL ({_date.today().isoformat()}) ***",
        )
    _write_cell_safe(ws, "B10", design.site_info.school_name)
    _write_cell_safe(ws, "B12", design.site_info.school_code)
    _write_cell_safe(ws, "B14", design.site_info.phone)
    _write_cell_safe(ws, "B18", design.site_info.install_tech)
    # Default to the worksheet's generation date when no install date was set
    # (parsed or entered) — but never clobber one the user already provided.
    install_date = design.site_info.install_date or date.today().strftime("%B %d %Y").upper()
    _write_cell_safe(ws, "B20", install_date)
    # B22 = "DMP XR550" (panel type — pre-filled in template)
    _write_cell_safe(ws, "B24", design.site_info.ip_address)
    # B25 = SUBNET MASK (pre-filled "255.255.255.0")
    _write_cell_safe(ws, "B26", design.site_info.default_gateway)

    # DMP XR550 sheet
    ws = wb["DMP XR550"]
    # B4 keeps template's pre-filled 'N/A' (the XR550 panel itself has no zone range).
    # D4 (merged D:E anchor) holds the panel location — writing to E4 would silently drop.
    _write_cell_safe(ws, "D4", design.site_info.xr550_location)

    # LX Bus 500/600/... sub-tables (rows 13-42). Clear all template defaults in B/C/D
    # (leave column A intact — it has the merged "LX Bus 500/600/..." group labels)
    # then write one row per RSP into the Bus 500 block.
    for r in range(13, 43):
        for col in ("B", "C", "D"):
            try:
                ws[f"{col}{r}"].value = None
            except Exception:
                pass
    for i, rsp in enumerate(design.rsps[:6]):
        row = 13 + i
        if rsp.zones:
            zmin, zmax = min(rsp.zones), max(rsp.zones)
            _write_cell_safe(ws, f"B{row}", f"{getattr(rsp, 'model', '') or '714-16'}-{rsp.number}")
            _write_cell_safe(ws, f"C{row}", f"{zmin}-{zmax}")
        _write_cell_safe(ws, f"D{row}", rsp.location)

    # DMP 714 Exp Mod sheet
    ws = wb["DMP 714 Exp Mod"]
    # Clear rows 4-40 first to remove template pre-fill
    for clear_row in range(4, 41):
        for col in ["A", "B", "C", "D", "E"]:
            try:
                ws[f"{col}{clear_row}"].value = None
            except:
                pass

    row_start = 4
    for i, rsp in enumerate(design.rsps):
        row = row_start + i
        model = getattr(rsp, "model", "") or "714-16"
        _write_cell_safe(ws, f"A{row}", f"DMP {model} #")
        _write_cell_safe(ws, f"B{row}", rsp.number)
        if rsp.zones:
            zone_min = min(rsp.zones)
            zone_max = max(rsp.zones)
            _write_cell_safe(ws, f"C{row}", f"{zone_min} - {zone_max}")
        _write_cell_safe(ws, f"D{row}", rsp.location)

    # Keypad sheet
    ws = wb["Keypad"]
    # Clear rows 3-30 first
    for clear_row in range(3, 31):
        for col in ["A", "B", "C", "D", "E"]:
            try:
                ws[f"{col}{clear_row}"].value = None
            except:
                pass

    row_start = 3
    for i, kp in enumerate(design.keypads):
        row = row_start + i
        _write_cell_safe(ws, f"A{row}", kp.number)
        _write_cell_safe(ws, f"B{row}", kp.source or "")
        _write_cell_safe(ws, f"C{row}", "Y" if kp.global_keypad else "N")
        _write_cell_safe(ws, f"D{row}", kp.location)

    # Resolve a splitter output description (e.g. 'RSP 2', 'KEYPAD #3', 'To LX 710-2', 'Spare')
    # to the destination device's location, so the splitter sheets can show *where* each
    # output cable runs to.
    def _resolve_output_location(output_desc: str) -> str:
        if not output_desc:
            return ""
        if output_desc == "Spare":
            return "Spare"
        # Accept both 'RSP 1' (legacy token) and 'RSP-1' (written convention).
        m = re.match(r"^RSP[\s-]*(\d+)$", output_desc, re.IGNORECASE)
        if m:
            n = int(m.group(1))
            for r in design.rsps:
                if r.number == n:
                    return r.location or ""
            return ""
        m = re.match(r"^KEYPAD\s*#?\s*(\d+)", output_desc, re.IGNORECASE)
        if m:
            n = int(m.group(1))
            for k in design.keypads:
                if k.number == n:
                    return k.location or ""
            return ""
        m = re.match(r"^To\s+(.+)$", output_desc, re.IGNORECASE)
        if m:
            target_id = m.group(1).strip()
            for s in design.splitters:
                if s.id == target_id:
                    return s.location or ""
            return ""
        return ""

    # 710 Splitter-Repeater(KP-Bus) sheet (note the trailing space in sheet name)
    ws = wb["710 Splitter-Repeater(KP-Bus) "] if "710 Splitter-Repeater(KP-Bus) " in wb.sheetnames else None
    if ws:
        # Clear rows 2-50 first
        for clear_row in range(2, 51):
            for col in ["A", "B", "C", "D", "E"]:
                try:
                    ws[f"{col}{clear_row}"].value = None
                except:
                    pass

        kp_splitters = [s for s in design.splitters if s.splitter_type == "KP"]
        row = 2
        for splitter in kp_splitters:
            _write_cell_safe(ws, f"A{row}", splitter.id)
            _write_cell_safe(ws, f"B{row}", "KP-Bus In")
            # Use the populated input description if Phase 3 (or interactive) provided one;
            # otherwise fall back to the canonical top-of-bus default.
            input_desc = splitter.inputs.get("KP-Bus In") or "KEYPAD BUS IN FROM XR/550"
            _write_cell_safe(ws, f"C{row}", input_desc)
            _write_cell_safe(ws, f"D{row}", splitter.location)
            row += 1
            # Write outputs if available, otherwise write 3 blank output rows
            outputs_to_write = splitter.outputs[:3] if splitter.outputs else ["", "", ""]
            for i, output in enumerate(outputs_to_write, 1):
                _write_cell_safe(ws, f"B{row}", f"KP-Bus {i}")
                if output:
                    _write_cell_safe(ws, f"C{row}", output)
                    dest_loc = _resolve_output_location(output)
                    if dest_loc:
                        _write_cell_safe(ws, f"D{row}", dest_loc)
                row += 1

    # 710 Splitter-Repeater LX500 sheet (check both "LX500" and "LXBus" name variants)
    ws_lx_name = None
    for name in wb.sheetnames:
        if "LX" in name and "710" in name:
            ws_lx_name = name
            break
    # If we found it but it's called "LXBus", rename it to "LX500" for compatibility with parse_dmp_worksheet
    if ws_lx_name and "LXBus" in ws_lx_name:
        ws = wb[ws_lx_name]
        ws.title = "710 Splitter-Repeater LX500"
        ws_lx_name = "710 Splitter-Repeater LX500"
    if ws_lx_name:
        ws = wb[ws_lx_name]
        # Clear rows 2-50 first
        for clear_row in range(2, 51):
            for col in ["A", "B", "C", "D", "E"]:
                try:
                    ws[f"{col}{clear_row}"].value = None
                except:
                    pass

        lx_splitters = [s for s in design.splitters if s.splitter_type == "LX"]
        row = 2
        for splitter in lx_splitters:
            _write_cell_safe(ws, f"A{row}", splitter.id)
            _write_cell_safe(ws, f"B{row}", "LX-Bus In")
            # Use the populated input description if Phase 3 (or interactive) provided one;
            # otherwise fall back to the canonical top-of-bus default.
            input_desc = splitter.inputs.get("LX-Bus In") or "500 BUS IN FROM XR/550"
            _write_cell_safe(ws, f"C{row}", input_desc)
            _write_cell_safe(ws, f"D{row}", splitter.location)
            row += 1
            # Write outputs if available, otherwise write 3 blank output rows
            outputs_to_write = splitter.outputs[:3] if splitter.outputs else ["", "", ""]
            for i, output in enumerate(outputs_to_write, 1):
                _write_cell_safe(ws, f"B{row}", f"LX-Bus {i}")
                if output:
                    _write_cell_safe(ws, f"C{row}", output)
                    dest_loc = _resolve_output_location(output)
                    if dest_loc:
                        _write_cell_safe(ws, f"D{row}", dest_loc)
                row += 1

    # DMP 505-12_G Power Supply 1-10 sheet
    ws = wb["DMP 505-12_G Power Supply 1-10"]
    # Clear rows 2-50 first
    for clear_row in range(2, 51):
        for col in ["A", "B", "C", "D", "E"]:
            try:
                ws[f"{col}{clear_row}"].value = None
            except:
                pass

    row = 2
    for ps in design.power_supplies:
        _write_cell_safe(ws, f"A{row}", ps.number)
        _write_cell_safe(ws, f"B{row}", "RELAY 1")
        _write_cell_safe(ws, f"C{row}", "12v DC Output to Terminal Strip")
        _write_cell_safe(ws, f"D{row}", ps.location)
        row += 1
        for relay_num in range(2, 5):
            _write_cell_safe(ws, f"B{row}", f"RELAY {relay_num}")
            _write_cell_safe(ws, f"C{row}", ps.relays.get(relay_num, ""))
            _write_cell_safe(ws, f"D{row}", ps.location)
            row += 1

    # Rename template's Point Info sheets to match the example's naming convention.
    # Template ships them as "DMP 714 Point Info 1"..."15"; example uses "DMP 714-16 Point Info (N)".
    # We do NOT touch cell contents — Point Info sheets are entirely formula-driven from Master,
    # so we inject zone descriptions into Master and let the formulas propagate.
    for name in list(wb.sheetnames):
        if "DMP 714" in name and "Point Info" in name and "714-16" not in name:
            m = re.search(r"Point Info\s+(\d+)", name)
            if m:
                wb[name].title = f"DMP 714-16 Point Info ({m.group(1)})"

    # Master sheet: column A holds zone labels (Z501, Z502, ..., Z981) and column B holds
    # zone descriptions. Build a zone_num -> master_row map by reading the existing labels
    # (Master has jumps at Z596->Z601 and Z696->Z701 for LX bus boundaries, so direct
    # arithmetic isn't safe — read the labels instead).
    #
    # IMPORTANT: column B has conditional formatting that highlights cells containing
    # "A/C LOSS" or "BATT. TRBL" (orange text). Supervisory zones must use those phrases
    # so the conditional formatting fires; writing a room name instead breaks the color.
    if "Master" in wb.sheetnames:
        from openpyxl.styles import Alignment
        master = wb["Master"]

        zone_to_row: dict[int, int] = {}
        for r in range(2, master.max_row + 1):
            v = master[f"A{r}"].value
            if isinstance(v, str) and v.startswith("Z"):
                try:
                    zone_to_row[int(v[1:])] = r
                except ValueError:
                    pass

        # Map each zone to its RSP and position-within-RSP. Within an RSP's 16-zone block,
        # the second-to-last zone is the A/C-loss supervisory and the last is BATT.
        zone_to_rsp_meta: dict[int, tuple[int, str]] = {}  # zone_num -> (rsp_num, "AC"|"BATT"|"NORMAL")
        for rsp in design.rsps:
            sorted_zones = sorted(rsp.zones)
            for idx, znum in enumerate(sorted_zones):
                if idx == len(sorted_zones) - 2:
                    zone_to_rsp_meta[znum] = (rsp.number, "AC")
                elif idx == len(sorted_zones) - 1:
                    zone_to_rsp_meta[znum] = (rsp.number, "BATT")
                else:
                    zone_to_rsp_meta[znum] = (rsp.number, "NORMAL")

        center = Alignment(horizontal="center", vertical="center")
        for zone in design.zones:
            row = zone_to_row.get(zone.number)
            if row is None:
                continue
            if zone.device_type == "Supervisory":
                rsp_num, ps_kind = zone_to_rsp_meta.get(zone.number, (0, "AC"))
                description = f"PS-{rsp_num}: {'A/C LOSS' if ps_kind == 'AC' else 'BATT. TRBL'}"
            elif zone.device_type == "Spare":
                description = "SPARE"
            else:
                description = zone.location or ""
            cell = master[f"B{row}"]
            cell.value = description
            # Re-apply explicit center alignment so the cell renders centered in Excel
            # regardless of whether the column-level style is honored.
            cell.alignment = center

        # Point Info detail sheets: the template wires each "Point Info (N)" to a FIXED
        # 16-row Master stride (module N -> Master rows 2+16(N-1)..). That only holds when
        # every module is a full 16-zone block; with packed numbering an 8-port module
        # shifts every later module off its stride. Rewire each sheet to its module's
        # ACTUAL zones: one data row (sheet rows 4..19) per zone, the last two flagged
        # Supervisory, unused rows blanked so we never bleed the next module's zones in.
        # (These edits survive _overlay_openpyxl_changes, which copies every worksheet XML.)
        rsp_by_num = {rsp.number: rsp for rsp in design.rsps}
        _PI_NUM_RE = re.compile(r"Point Info\s*\(?\s*(\d+)")
        for sheet_name in wb.sheetnames:
            if "Point Info" not in sheet_name:
                continue
            m = _PI_NUM_RE.search(sheet_name)
            if not m:
                continue
            pi = wb[sheet_name]
            rsp = rsp_by_num.get(int(m.group(1)))
            sorted_zones = sorted(rsp.zones) if rsp else []
            model = (getattr(rsp, "model", "") or "714-16") if rsp else "714-16"
            for i in range(16):  # template data rows 4..19; row 20 is the "Source:" label
                row = 4 + i
                if i < len(sorted_zones):
                    mrow = zone_to_row.get(sorted_zones[i])
                    pi[f"A{row}"] = f"=Master!A{mrow}" if mrow else None
                    pi[f"B{row}"] = f"=Master!B{mrow}" if mrow else None
                    pi[f"D{row}"] = 1
                    pi[f"F{row}"] = "Supervisory" if i >= len(sorted_zones) - 2 else "Motion"
                else:
                    pi[f"A{row}"] = None
                    pi[f"B{row}"] = None
                    pi[f"D{row}"] = None
                    pi[f"F{row}"] = None
            if rsp:
                pi["A20"] = f"Source: DMP {model} Expander #{rsp.number}"

    # Write the school address as workbook custom doc properties — invisible in
    # Excel's sheet UI but readable by parse_dmp_worksheet via wb.custom_doc_props.
    # The door chart's Header B4/B5 populate from these.
    from openpyxl.packaging.custom import StringProperty
    if design.site_info.address_line1:
        wb.custom_doc_props.append(
            StringProperty(name="SchoolAddressLine1", value=design.site_info.address_line1)
        )
    if design.site_info.address_line2:
        wb.custom_doc_props.append(
            StringProperty(name="SchoolAddressLine2", value=design.site_info.address_line2)
        )
    if stamp:
        # DRAFT exports are working copies of a session, never inputs; the
        # app's import path refuses workbooks carrying DMPStatus=DRAFT.
        wb.custom_doc_props.append(StringProperty(name="DMPStatus", value=stamp))

    # openpyxl strips files Excel needs (queryTables, connections, calcChain, sharedStrings,
    # printerSettings, table _rels, customXml, etc.) and reformats others (Content_Types,
    # workbook.xml.rels) in ways that break Excel's "External data range" repair.
    # Strategy: save openpyxl output to a temp file, then BINARY COPY the template to the
    # final path (perfect formatting), then overlay only the files openpyxl actually modified
    # (worksheets, styles, workbook, workbook rels). Everything else stays from template.
    import shutil
    import zipfile
    from tempfile import NamedTemporaryFile

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmpf:
        openpyxl_tmp_path = Path(tmpf.name)
    try:
        wb.save(openpyxl_tmp_path)
        _overlay_openpyxl_changes(template_path, openpyxl_tmp_path, output_path)
    finally:
        if openpyxl_tmp_path.exists():
            openpyxl_tmp_path.unlink()

    # Final post-processing on the now-template-based output:
    #   - Restore conditional formatting blocks (openpyxl mangled the dxfId references in
    #     the worksheet XML we just overlaid)
    #   - Restore Master sheet's worksheet header (namespace decls dropped by openpyxl)
    #   - Strip calcChain.xml — it's a calculation-order cache that's now stale because
    #     our overlaid worksheets changed cell content. Excel rebuilds it on open.
    _restore_conditional_formatting(template_path, output_path)
    _restore_master_sheet_header(template_path, output_path)
    _strip_calc_chain(output_path)

    # Trim the template's 15 Point Info sheets down to one per RSP / 714-16 expander.
    # Keep sheets up to the highest module number — Point Info sheet N is
    # hard-wired to module N's Master row stride, so numbering gaps (a removed
    # expander) must keep their (blank) sheet rather than shift later modules.
    _remove_extra_point_info_sheets(
        output_path, keep_count=max((r.number for r in design.rsps), default=0))


def _strip_calc_chain(output_path: Path) -> None:
    """Remove xl/calcChain.xml and all references to it. calcChain is a cached calculation
    order; after we've overlaid worksheet content, Excel detects the cache as stale and
    flags it on open ("Removed Records: Formula from /xl/calcChain.xml part"). Removing it
    cleanly causes Excel to rebuild the cache silently on open.
    """
    import re
    import shutil
    import zipfile
    from tempfile import NamedTemporaryFile

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmpf:
        tmp_path = tmpf.name

    with zipfile.ZipFile(output_path, "r") as zin, zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.namelist():
            if item == "xl/calcChain.xml":
                continue  # drop the cache file
            data = zin.read(item)
            if item == "[Content_Types].xml":
                # Remove the calcChain Override declaration
                text = data.decode("utf-8")
                text = re.sub(
                    r'<Override\s+PartName="/xl/calcChain\.xml"[^>]*?/>',
                    "",
                    text,
                )
                data = text.encode("utf-8")
            elif item == "xl/_rels/workbook.xml.rels":
                # Remove the calcChain Relationship
                text = data.decode("utf-8")
                text = re.sub(
                    r'<Relationship\s+[^>]*?Type="[^"]*calcChain"[^>]*?/>',
                    "",
                    text,
                )
                data = text.encode("utf-8")
            zout.writestr(item, data)
    shutil.move(tmp_path, output_path)


def _remove_extra_point_info_sheets(output_path: Path, keep_count: int) -> None:
    """Delete unused 'Point Info' sheets so the workbook keeps exactly `keep_count`
    of them (one per RSP / 714-16 expander module).

    The template ships 15 Point Info sheets; real designs use far fewer. They are
    the LAST sheets in the workbook, so only trailing parts are removed — no rId
    renumbering of earlier sheets is needed. Done as a post-process on the final
    zip because _overlay_openpyxl_changes() rebuilds the output from a binary copy
    of the template (which still contains all 15 Point Info sheets).
    """
    import re
    import shutil
    import zipfile
    from tempfile import NamedTemporaryFile

    if keep_count < 0:
        keep_count = 0

    with zipfile.ZipFile(output_path, "r") as zin:
        workbook_xml = zin.read("xl/workbook.xml").decode("utf-8")
        rels_xml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")

    # Collect the Point Info <sheet> entries, ordered by their trailing number.
    point_info: list[tuple[int, str, str]] = []  # (number, sheet_tag, r:id)
    for m in re.finditer(r"<sheet\b[^>]*?/>", workbook_xml):
        tag = m.group(0)
        name_m = re.search(r'name="([^"]*)"', tag)
        if not name_m or "Point Info" not in name_m.group(1):
            continue
        num_m = re.search(r"Point Info\s*\(?\s*(\d+)", name_m.group(1))
        rid_m = re.search(r'r:id="([^"]*)"', tag)
        if num_m and rid_m:
            point_info.append((int(num_m.group(1)), tag, rid_m.group(1)))
    point_info.sort(key=lambda t: t[0])

    to_delete = point_info[keep_count:]
    if not to_delete:
        return

    del_tags = [tag for _, tag, _ in to_delete]
    del_rids = {rid for _, _, rid in to_delete}

    # Map each deleted r:id to its worksheet part via workbook.xml.rels.
    del_parts: set[str] = set()
    for rid in del_rids:
        rel_m = re.search(
            r'<Relationship\b[^>]*?\bId="' + re.escape(rid) + r'"[^>]*?/>', rels_xml
        )
        if not rel_m:
            continue
        target_m = re.search(r'Target="([^"]*)"', rel_m.group(0))
        if not target_m:
            continue
        part = target_m.group(1).lstrip("/")
        if not part.startswith("xl/"):
            part = "xl/" + part
        del_parts.add(part)

    # Per-sheet _rels files for the deleted parts (xl/worksheets/_rels/sheetN.xml.rels).
    del_rel_parts: set[str] = set()
    for part in del_parts:
        head, tail = part.rsplit("/", 1)
        del_rel_parts.add(f"{head}/_rels/{tail}.rels")

    # workbook.xml: drop the deleted <sheet> elements.
    new_workbook = workbook_xml
    for tag in del_tags:
        new_workbook = new_workbook.replace(tag, "")

    # workbook.xml.rels: drop the deleted <Relationship> elements.
    new_rels = rels_xml
    for rid in del_rids:
        new_rels = re.sub(
            r'<Relationship\b[^>]*?\bId="' + re.escape(rid) + r'"[^>]*?/>', "", new_rels
        )

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmpf:
        tmp_path = tmpf.name

    with zipfile.ZipFile(output_path, "r") as zin, zipfile.ZipFile(
        tmp_path, "w", zipfile.ZIP_DEFLATED
    ) as zout:
        for item in zin.namelist():
            if item in del_parts or item in del_rel_parts:
                continue  # drop deleted worksheet parts and their _rels
            data = zin.read(item)
            if item == "xl/workbook.xml":
                data = new_workbook.encode("utf-8")
            elif item == "xl/_rels/workbook.xml.rels":
                data = new_rels.encode("utf-8")
            elif item == "[Content_Types].xml":
                text = data.decode("utf-8")
                for part in del_parts:
                    text = re.sub(
                        r'<Override\b[^>]*?PartName="/' + re.escape(part) + r'"[^>]*?/>',
                        "",
                        text,
                    )
                data = text.encode("utf-8")
            zout.writestr(item, data)
    shutil.move(tmp_path, output_path)


def _overlay_openpyxl_changes(template_path: Path, openpyxl_tmp_path: Path, output_path: Path) -> None:
    """Build output by starting from the template (binary copy) and overlaying only the
    files openpyxl actually modified.

    openpyxl's save drops files Excel needs (queryTables, connections, calcChain,
    sharedStrings, printerSettings, table _rels, customXml) and reformats Content_Types,
    breaking Excel's parts model. By using template as the base, we keep all those parts
    intact. Only worksheets, styles, workbook, and workbook rels need to come from
    openpyxl's output (those carry our cell changes, sheet renames, and any new styles).

    The rId mapping in workbook.xml is preserved across template and openpyxl save (both
    use rId1..rIdN in sheet order), so workbook.xml from openpyxl is compatible with the
    template's workbook.xml.rels.
    """
    import re as _re
    import shutil
    import zipfile

    OVERLAY_PREFIXES = ("xl/worksheets/sheet",)
    OVERLAY_FILES = {"xl/styles.xml", "xl/workbook.xml"}
    # New parts that openpyxl may add (and the template lacks). For each, we copy
    # the part verbatim and patch [Content_Types].xml + _rels/.rels to declare it.
    NEW_PARTS = {"docProps/custom.xml"}

    # Read openpyxl's modified files
    overlays: dict[str, bytes] = {}
    new_parts: dict[str, bytes] = {}
    with zipfile.ZipFile(openpyxl_tmp_path) as zop:
        names = set(zop.namelist())
        for name in names:
            if name in OVERLAY_FILES or any(
                name.startswith(p) and name.endswith(".xml") and "_rels" not in name
                for p in OVERLAY_PREFIXES
            ):
                overlays[name] = zop.read(name)
            elif name in NEW_PARTS:
                new_parts[name] = zop.read(name)

    # Start from the template (binary copy preserves everything Excel needs)
    shutil.copy(template_path, output_path)

    # Overlay changed files onto the template-based output
    from tempfile import NamedTemporaryFile

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmpf:
        rebuild_path = tmpf.name
    with zipfile.ZipFile(output_path, "r") as zin, zipfile.ZipFile(rebuild_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.namelist():
            if item == "[Content_Types].xml" and "docProps/custom.xml" in new_parts:
                # Inject the custom-properties Override declaration if missing
                ct_xml = zin.read(item).decode("utf-8")
                if "docProps/custom.xml" not in ct_xml:
                    override = (
                        '<Override PartName="/docProps/custom.xml" '
                        'ContentType="application/vnd.openxmlformats-officedocument.custom-properties+xml"/>'
                    )
                    ct_xml = ct_xml.replace("</Types>", override + "</Types>")
                zout.writestr(item, ct_xml.encode("utf-8"))
            elif item == "_rels/.rels" and "docProps/custom.xml" in new_parts:
                # Inject the custom-properties Relationship if missing, choosing an unused rId
                rels_xml = zin.read(item).decode("utf-8")
                if "docProps/custom.xml" not in rels_xml:
                    used_ids = set(_re.findall(r'Id="(rId\d+)"', rels_xml))
                    next_id = 1
                    while f"rId{next_id}" in used_ids:
                        next_id += 1
                    rel = (
                        f'<Relationship Id="rId{next_id}" '
                        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/custom-properties" '
                        'Target="docProps/custom.xml"/>'
                    )
                    rels_xml = rels_xml.replace("</Relationships>", rel + "</Relationships>")
                zout.writestr(item, rels_xml.encode("utf-8"))
            else:
                data = overlays.get(item, zin.read(item))
                zout.writestr(item, data)
        # Append new parts that the template didn't have
        for name, data in new_parts.items():
            zout.writestr(name, data)
    shutil.move(rebuild_path, output_path)


def _restore_master_sheet_header(template_path: Path, output_path: Path) -> None:
    """Replace the Master sheet's worksheet root element + everything up to <sheetData> with
    the template's, preserving namespace declarations (xmlns:r, xmlns:x14ac, etc.) that
    openpyxl drops on save. Without xmlns:r at the worksheet root, Excel cannot bind the
    <tablePart r:id="rId1"/> element to xl/tables/table1.xml, which breaks the table-style
    row banding (TableStyleMedium7).

    Cell data inside <sheetData> is left untouched — that's where our zone descriptions live.
    """
    import shutil
    import zipfile
    from tempfile import NamedTemporaryFile

    def _find_master_sheet(zip_path: Path) -> tuple[str, str] | None:
        with zipfile.ZipFile(zip_path) as z:
            for name in z.namelist():
                if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                    xml = z.read(name).decode("utf-8")
                    # Master sheet is the only worksheet that has both <conditionalFormatting>
                    # blocks and the table reference
                    if "<conditionalFormatting" in xml and "tablePart" in xml:
                        return name, xml
        return None

    tpl_match = _find_master_sheet(template_path)
    out_match = _find_master_sheet(output_path)
    if not tpl_match or not out_match:
        return

    tpl_xml = tpl_match[1]
    target_name, out_xml = out_match

    # Extract template's header (everything before <sheetData>) and output's body (sheetData onwards).
    tpl_split = tpl_xml.find("<sheetData")
    out_split = out_xml.find("<sheetData")
    if tpl_split < 0 or out_split < 0:
        return

    new_xml = tpl_xml[:tpl_split] + out_xml[out_split:]

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmpf:
        tmp_path = tmpf.name
    with zipfile.ZipFile(output_path, "r") as zin, zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.namelist():
            data = new_xml.encode("utf-8") if item == target_name else zin.read(item)
            zout.writestr(item, data)
    shutil.move(tmp_path, output_path)


def _restore_template_files(template_path: Path, output_path: Path, file_paths: list[str]) -> None:
    """Copy specific files byte-for-byte from template's zip into the output's zip.

    Use for files that openpyxl mangles or simplifies on save in ways Excel cares about
    (e.g. table definitions losing xr:uid revision-tracking attributes).
    """
    import shutil
    import zipfile
    from tempfile import NamedTemporaryFile

    with zipfile.ZipFile(template_path) as ztpl:
        tpl_files: dict[str, bytes] = {}
        for fp in file_paths:
            try:
                tpl_files[fp] = ztpl.read(fp)
            except KeyError:
                pass

    if not tpl_files:
        return

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmpf:
        tmp_path = tmpf.name
    with zipfile.ZipFile(output_path, "r") as zin, zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.namelist():
            data = tpl_files.get(item) if item in tpl_files else zin.read(item)
            zout.writestr(item, data)
    shutil.move(tmp_path, output_path)


def _restore_conditional_formatting(template_path: Path, output_path: Path) -> None:
    """Copy <conditionalFormatting> blocks from template's Master sheet into the saved output.

    openpyxl's save process reassigns every cfRule's dxfId to 0 even though the dxfs
    collection in styles.xml is preserved. The cell positions on the Master sheet don't
    move between template and output, so a literal substitution restores the rules
    intact without disturbing anything else.
    """
    import re
    import shutil
    import zipfile
    from tempfile import NamedTemporaryFile

    cf_pattern = re.compile(
        r"<conditionalFormatting\b[^>]*>.*?</conditionalFormatting>",
        re.DOTALL,
    )

    # The Master sheet is the only worksheet with <conditionalFormatting> blocks —
    # other sheets only show cached formula values like "PS-1: A/C LOSS". Use the
    # presence of the cfRules-containing element itself to identify it in both files.
    def _find_cf_sheet(zip_path: Path) -> tuple[str, str] | None:
        with zipfile.ZipFile(zip_path) as z:
            for name in z.namelist():
                if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                    xml = z.read(name).decode("utf-8")
                    if "<conditionalFormatting" in xml and "A/C LOSS" in xml:
                        return name, xml
        return None

    tpl_match = _find_cf_sheet(template_path)
    out_match = _find_cf_sheet(output_path)
    if not tpl_match or not out_match:
        return
    tpl_sheet_xml = tpl_match[1]
    target_name, out_sheet_xml = out_match

    original_cf_blocks = cf_pattern.findall(tpl_sheet_xml)
    if not original_cf_blocks:
        return

    # Replace all openpyxl-mangled <conditionalFormatting> blocks with the originals.
    # Splice them in at the location of the first existing block (preserves ordering
    # relative to other elements like <pageMargins>, <printOptions>).
    matches = list(cf_pattern.finditer(out_sheet_xml))
    if not matches:
        return
    insert_at = matches[0].start()
    end_at = matches[-1].end()
    new_xml = (
        out_sheet_xml[:insert_at]
        + "".join(original_cf_blocks)
        + out_sheet_xml[end_at:]
    )

    # Write the patched xml back to the output zip.
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmpf:
        tmp_path = tmpf.name
    with zipfile.ZipFile(output_path, "r") as zin, zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.namelist():
            data = new_xml.encode("utf-8") if item == target_name else zin.read(item)
            zout.writestr(item, data)
    shutil.move(tmp_path, output_path)


def main():
    ap = argparse.ArgumentParser(description="Generate DMP Installation Worksheet from design PDF.")
    ap.add_argument("pdf", help="Path to the design PDF.")
    ap.add_argument(
        "--searchable",
        default=None,
        help="Path to an already-OCR'd searchable PDF (skips ocrmypdf if provided).",
    )
    ap.add_argument(
        "--output",
        default=None,
        help="Output path for the DMP worksheet. Default: output/<school>_dmp_<date>.xlsx",
    )
    ap.add_argument(
        "--non-interactive",
        action="store_true",
        help="Skip interactive prompts; leave metadata gaps blank/defaulted.",
    )
    ap.add_argument(
        "--prompt-routing",
        action="store_true",
        help="Prompt for splitter I/O routing (overrides auto-derived defaults).",
    )
    args = ap.parse_args()

    pdf_path = Path(args.pdf).resolve()
    if not pdf_path.exists():
        sys.exit(f"PDF not found: {pdf_path}")

    template_path = DEFAULT_TEMPLATE
    if not template_path.exists():
        sys.exit(f"Template not found: {template_path}")

    print(f"[1/4] Ensuring searchable PDF...")
    searchable_pdf = ensure_searchable_pdf(
        pdf_path,
        Path(args.searchable) if args.searchable else None,
    )

    # Topology extraction needs the ORIGINAL (pre-OCR) PDF: ocrmypdf can mangle
    # splitter labels on the riser, making them invisible on the searchable copy.
    original_pdf = resolve_original_pdf(pdf_path)
    if original_pdf != pdf_path:
        print(f"      Using {original_pdf.name} for topology (original preserves vector text).")

    print(f"[2/4] Parsing zone schedule and topology...")
    design = build_dmp_design_from_pdf(
        searchable_pdf, original_pdf, args.non_interactive, args.prompt_routing
    )

    # Determine output path
    if args.output:
        output_path = Path(args.output).resolve()
    else:
        school_slug = _slugify(design.site_info.school_name or "output")
        output_dir = DEFAULT_OUTPUT_DIR
        output_dir.mkdir(exist_ok=True)
        out_name = f"{school_slug}_dmp_{date.today().isoformat()}.xlsx"
        output_path = output_dir / out_name

    print(f"[3/4] Writing DMP worksheet...")
    write_dmp_xlsx(design, template_path, output_path)

    print(f"[4/4] Done!")
    print(f"Output: {output_path}")


if __name__ == "__main__":
    main()
