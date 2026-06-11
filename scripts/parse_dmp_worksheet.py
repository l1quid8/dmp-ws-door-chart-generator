"""
Parses DMP worksheet (.xlsx) to extract site info, splitters, RSPs, keypads, and terminal descriptions.

The DMP worksheet contains structured tabs:
  - SITE INFO: school name, code, contact info
  - DMP XR550: location of main panel + LX bus info
  - 710 Splitter-Repeater(KP-Bus): Keypad splitters (KP-710-N) with I/O mappings
  - 710 Splitter-Repeater LX500: LX 710-N splitters with I/O mappings
  - Keypad: list of keypads and their source (MSP or KP splitter)
  - DMP 714 Exp Mod: expansion modules and their point ranges
  - DMP 714-16 Point Info (1-4): detailed zone descriptions per expansion module

Output: DMPDesign object with structured data for injection into the Master sheet.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl


# -------- data model --------

@dataclass
class SiteInfo:
    school_name: Optional[str] = None
    school_code: Optional[str] = None
    phone: Optional[str] = None
    install_tech: Optional[str] = None
    install_date: Optional[str] = None
    ip_address: Optional[str] = None
    default_gateway: Optional[str] = None
    xr550_location: Optional[str] = None  # from DMP XR550 sheet
    # Stored as workbook custom doc properties (invisible in Excel sheet UI)
    address_line1: Optional[str] = None
    address_line2: Optional[str] = None


@dataclass
class Splitter:
    """A 710-bus splitter (LX or KP)."""
    id: str                        # "LX 710-1", "KP-710-1"
    splitter_type: str             # "LX" or "KP"
    location: Optional[str] = None
    inputs: dict[str, str] = field(default_factory=dict)  # "LX-Bus In" / "KP-Bus In" -> description
    outputs: list[str] = field(default_factory=list)      # ["RSP 1", "RSP 2", ...]


@dataclass
class RSP:
    """Remote Service Panel."""
    number: int
    location: Optional[str] = None
    zones: list[int] = field(default_factory=list)


@dataclass
class Keypad:
    """Keypad module."""
    number: int
    source: Optional[str] = None      # "MSP" or "710-KP-1"
    location: Optional[str] = None
    global_keypad: bool = False


@dataclass
class PowerSupply:
    """Power Supply module (DMP 505-12G)."""
    number: int
    location: Optional[str] = None
    relays: dict[int, str] = field(default_factory=dict)  # relay number -> function


@dataclass
class ZoneInfo:
    """Zone from the point-info sheets."""
    number: int
    location: Optional[str] = None
    device_type: Optional[str] = None  # "Motion", "Supervisory", etc.
    partition: Optional[int] = None


@dataclass
class Zone:
    """A zone row from the DMP's Master sheet (the door chart's source for zone descriptions).

    Distinct from ZoneInfo (which is sourced from the per-RSP point-info sheets).
    """
    number: int                              # e.g. 501
    description: str                         # e.g. 'CLASSROOM 1', 'SPARE', 'PS-2: A/C LOSS'
    rsp_number: Optional[int] = None         # 1..N for normal/PS rows; None for SPARE
    is_spare: bool = False
    is_ps_ac: bool = False
    is_ps_batt: bool = False


@dataclass
class DMPDesign:
    site_info: SiteInfo = field(default_factory=SiteInfo)
    splitters: list[Splitter] = field(default_factory=list)
    rsps: list[RSP] = field(default_factory=list)
    keypads: list[Keypad] = field(default_factory=list)
    power_supplies: list[PowerSupply] = field(default_factory=list)
    zones: list[ZoneInfo] = field(default_factory=list)
    master_zones: list[Zone] = field(default_factory=list)   # NEW: from Master sheet
    conflicts: list = field(default_factory=list)            # unresolved source-data conflicts
    topology_source: str = ""                                # "riser" | "auto-derived"
    master_zones_source: str = ""                            # "master" | "point_info"


# -------- parsing helpers --------

def parse_dmp_worksheet(xlsx_path: str | Path) -> DMPDesign:
    """Top-level: parse the DMP worksheet and return structured data."""
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    design = DMPDesign()

    # Parse each sheet
    if "SITE INFO" in wb.sheetnames:
        design.site_info = _parse_site_info(wb["SITE INFO"])

    # School address lives in workbook custom doc properties (invisible in Excel UI).
    # Populated by generate_dmp_ws.py from the PDF's title block.
    for prop in getattr(wb.custom_doc_props, "props", []):
        if prop.name == "SchoolAddressLine1":
            design.site_info.address_line1 = prop.value
        elif prop.name == "SchoolAddressLine2":
            design.site_info.address_line2 = prop.value

    if "DMP XR550" in wb.sheetnames:
        design.site_info.xr550_location = _parse_xr550_location(wb["DMP XR550"])

    if "710 Splitter-Repeater(KP-Bus) " in wb.sheetnames:
        kp_splitters = _parse_kp_splitters(wb["710 Splitter-Repeater(KP-Bus) "])
        design.splitters.extend(kp_splitters)

    if "710 Splitter-Repeater LX500" in wb.sheetnames:
        lx_splitters = _parse_lx_splitters(wb["710 Splitter-Repeater LX500"])
        design.splitters.extend(lx_splitters)

    if "Keypad" in wb.sheetnames:
        design.keypads = _parse_keypads(wb["Keypad"])

    # 714-16 / 714-08 expansion modules ARE the RSPs — module# = RSP#, point range = zones
    for sheet_name in wb.sheetnames:
        if sheet_name.strip() == "DMP 714 Exp Mod":
            design.rsps = _parse_rsps(wb[sheet_name])
            break

    if "Master" in wb.sheetnames:
        design.master_zones = _parse_master_zones(wb["Master"], design.rsps)

    if "DMP 505-12_G Power Supply 1-10" in wb.sheetnames:
        design.power_supplies = _parse_power_supplies(wb["DMP 505-12_G Power Supply 1-10"])

    # Parse all point-info sheets
    for sheet_name in wb.sheetnames:
        if "DMP 714-16 Point Info" in sheet_name:
            zones = _parse_point_info(wb[sheet_name])
            design.zones.extend(zones)

    # Older worksheets predate the Master sheet but still carry every zone in the
    # Point Info sheets (parsed into design.zones above). When Master is absent,
    # reconstruct master_zones from that data so the door chart's zone area still
    # populates. New-format worksheets keep a Master sheet and skip this entirely.
    design.master_zones_source = "master" if design.master_zones else ""
    if not design.master_zones and design.zones:
        design.master_zones = _master_zones_from_point_info(design.zones, design.rsps)
        design.master_zones_source = "point_info"
        print(f"  Master sheet absent — reconstructed {len(design.master_zones)} "
              f"zones from Point Info sheets")

    wb.close()
    return design


def worksheet_looks_like_dmp(design: DMPDesign) -> bool:
    """True if a parsed worksheet carries the load-bearing DMP signals.

    parse_dmp_worksheet is permissive — every sheet is read behind an
    ``if NAME in wb.sheetnames`` guard, so a wrong/unrelated .xlsx parses into an
    (almost) empty DMPDesign instead of raising. inject() is likewise defensive and
    would then silently write a *blank* door chart. Callers that accept an arbitrary
    user-supplied workbook should gate on this: require at least a school name
    (from SITE INFO) or zone rows (from Master) before treating it as a real DMP
    worksheet.
    """
    return bool(design.site_info.school_name or design.master_zones)


def _parse_site_info(ws) -> SiteInfo:
    """Extract school name, code, phone from SITE INFO sheet."""
    info = SiteInfo()
    for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
        if not row or not row[0]:
            continue
        label = str(row[0]).strip()
        value = row[1] if len(row) > 1 else None
        if not value:
            continue

        if "School Name" in label:
            info.school_name = str(value).strip()
        elif "School Code" in label:
            info.school_code = str(value).strip()
        elif "Main Phone" in label:
            info.phone = str(value).strip()
        elif "Install Tech" in label:
            info.install_tech = str(value).strip()

    return info


def _parse_xr550_location(ws) -> Optional[str]:
    """Extract XR550 location from DMP XR550 sheet (cell D4)."""
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        if not row:
            continue
        # The "DMP XR550  #" row carries the panel location in column D (index 3).
        if row[0] and "DMP XR550" in str(row[0]):
            location = row[3] if len(row) > 3 else None
            if location:
                return str(location).strip()
    return None


def _parse_kp_splitters(ws) -> list[Splitter]:
    """Extract KP splitters from '710 Splitter-Repeater(KP-Bus) ' sheet."""
    splitters: list[Splitter] = []
    current_splitter: Optional[Splitter] = None

    for row in ws.iter_rows(min_row=1, max_row=50, values_only=True):
        if not row:
            continue

        id_text = str(row[0]).strip() if row[0] else ""
        func_text = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        desc_text = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        loc_text = str(row[3]).strip() if len(row) > 3 and row[3] else ""

        # Device ID line — accept legacy ("KP-710-1", "KP 710-1") and IA-diagram ("710-KP-1") formats
        if id_text.startswith("KP-") or id_text.startswith("KP ") or id_text.startswith("710-KP"):
            splitter_id = id_text.replace(" ", "-").upper()
            if current_splitter:
                splitters.append(current_splitter)
            current_splitter = Splitter(
                id=splitter_id,
                splitter_type="KP",
                location=loc_text if loc_text else None,
            )
            # First line may also have input info
            if func_text and "In" in func_text:
                current_splitter.inputs[func_text] = desc_text
        # I/O line (e.g., "KP-Bus In", "KP-Bus 1", etc.) — col A is empty, col B has function
        elif current_splitter and func_text:
            # Store as input or output based on "In" vs numbered
            if "In" in func_text:
                current_splitter.inputs[func_text] = desc_text
            else:
                # Output line: store the description (keypad or device it feeds)
                current_splitter.outputs.append(desc_text)

    if current_splitter:
        splitters.append(current_splitter)

    return splitters


def _parse_lx_splitters(ws) -> list[Splitter]:
    """Extract LX splitters from '710 Splitter-Repeater LX500' sheet."""
    splitters: list[Splitter] = []
    current_splitter: Optional[Splitter] = None

    for row in ws.iter_rows(min_row=1, max_row=50, values_only=True):
        if not row:
            continue

        id_text = str(row[0]).strip() if row[0] else ""
        func_text = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        desc_text = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        loc_text = str(row[3]).strip() if len(row) > 3 and row[3] else ""

        # Device ID line — accept legacy ("LX-710-1", "LX 710-1") and IA-diagram ("710-LX500-1") formats
        if id_text.startswith("LX ") or id_text.startswith("LX-") or id_text.startswith("710-LX"):
            splitter_id = id_text.replace(" ", "-").upper()
            if current_splitter:
                splitters.append(current_splitter)
            current_splitter = Splitter(
                id=splitter_id,
                splitter_type="LX",
                location=loc_text if loc_text else None,
            )
            # First line may also have input info
            if func_text and "In" in func_text:
                current_splitter.inputs[func_text] = desc_text
        # I/O line (e.g., "LX-Bus In", "LX-Bus 1", etc.) — col A is empty, col B has function
        elif current_splitter and func_text:
            if "In" in func_text:
                current_splitter.inputs[func_text] = desc_text
            else:
                # Output line: store the description
                current_splitter.outputs.append(desc_text)

    if current_splitter:
        splitters.append(current_splitter)

    return splitters


def _parse_rsps(ws) -> list[RSP]:
    """Extract RSPs from 'DMP 714 Exp Mod' sheet.

    Each row of the form `DMP 714-16 # | <num> | <range> | <location>` corresponds
    to one expansion module — and in C1's design convention, each expansion
    module IS an RSP (RSP1=module1, RSP2=module2, ...).

    Point range column may be a string like "501 - 516" — parse into a zone list.
    """
    rsps: list[RSP] = []
    for row in ws.iter_rows(min_row=4, max_row=40, values_only=True):
        if not row or not row[0]:
            continue
        prefix = str(row[0]).strip()
        if not prefix.startswith("DMP 714"):
            continue

        # Module number — col B (index 1)
        num_raw = row[1] if len(row) > 1 else None
        try:
            module_num = int(num_raw) if num_raw is not None else None
        except (ValueError, TypeError):
            module_num = None
        if module_num is None:
            continue

        # Point range — col C (index 2), e.g. "501 - 516"
        zones: list[int] = []
        if len(row) > 2 and row[2]:
            range_text = str(row[2]).strip()
            if "-" in range_text:
                parts = [p.strip() for p in range_text.split("-")]
                if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
                    lo, hi = int(parts[0]), int(parts[1])
                    if lo <= hi:
                        zones = list(range(lo, hi + 1))

        # Location — col D (index 3)
        location = None
        if len(row) > 3 and row[3]:
            location = str(row[3]).strip()

        rsps.append(RSP(number=module_num, location=location, zones=zones))
    return rsps


def _parse_keypads(ws) -> list[Keypad]:
    """Extract keypad modules from Keypad sheet."""
    keypads: list[Keypad] = []
    for row in ws.iter_rows(min_row=3, max_row=30, values_only=True):
        if not row or not row[0]:
            continue

        number_text = str(row[0]).strip() if row[0] else ""
        if not number_text.isdigit():
            continue

        kp_num = int(number_text)
        source = str(row[1]).strip() if len(row) > 1 and row[1] else None
        global_flag = row[2] if len(row) > 2 else None
        location = str(row[3]).strip() if len(row) > 3 and row[3] else None

        keypads.append(Keypad(
            number=kp_num,
            source=source,
            location=location,
            global_keypad=str(global_flag).upper() == "Y" if global_flag else False,
        ))

    return keypads


def _parse_power_supplies(ws) -> list[PowerSupply]:
    """Extract power supplies from DMP 505-12_G sheet."""
    power_supplies: list[PowerSupply] = []
    current_ps: Optional[PowerSupply] = None

    for row in ws.iter_rows(min_row=2, max_row=30, values_only=True):
        if not row:
            continue

        ps_num_text = str(row[0]).strip() if row[0] else ""
        func_text = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        desc_text = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        loc_text = str(row[3]).strip() if len(row) > 3 and row[3] else ""

        # New PS line
        if ps_num_text.isdigit():
            ps_num = int(ps_num_text)
            if current_ps:
                power_supplies.append(current_ps)
            current_ps = PowerSupply(number=ps_num, location=loc_text if loc_text else None)
        # Relay line
        elif current_ps and "RELAY" in func_text:
            relay_match = func_text.split()
            if len(relay_match) >= 2 and relay_match[1].isdigit():
                relay_num = int(relay_match[1])
                current_ps.relays[relay_num] = desc_text

    if current_ps:
        power_supplies.append(current_ps)

    return power_supplies


def _parse_point_info(ws) -> list[ZoneInfo]:
    """Extract zone info from a DMP 714-16 Point Info sheet."""
    zones: list[ZoneInfo] = []
    for row in ws.iter_rows(min_row=4, max_row=20, values_only=True):
        if not row or not row[1]:
            continue

        z_prefix = row[0]  # "Z"
        zone_num_text = str(row[1]).strip() if row[1] else ""
        location_text = str(row[2]).strip() if len(row) > 2 and row[2] else None
        device_type = str(row[6]).strip() if len(row) > 6 and row[6] else None

        if not zone_num_text.isdigit():
            continue

        zone_num = int(zone_num_text)
        zones.append(ZoneInfo(
            number=zone_num,
            location=location_text,
            device_type=device_type,
        ))

    return zones


import re as _re_for_zones  # noqa: E402 — placed here to keep top-of-file imports minimal

_PS_AC_RE = _re_for_zones.compile(r"^PS-(\d+):\s*A/?C", _re_for_zones.IGNORECASE)
_PS_BATT_RE = _re_for_zones.compile(r"^PS-(\d+):\s*BATT", _re_for_zones.IGNORECASE)


def _parse_master_zones(ws, rsps: list[RSP]) -> list[Zone]:
    """Read zone rows from the DMP's Master sheet.

    Layout: A1='ZONE #', B1='ZONE DESCRIPTION', then zone rows from row 2:
      A col: 'Z501'..'Z980'
      B col: room name / 'SPARE' / 'PS-N: A/C LOSS' / 'PS-N: BATT. TRBL'

    rsp_number for non-supervisory zones is derived from the RSP zone ranges
    (which were already parsed from 'DMP 714 Exp Mod'). For PS rows it comes
    from the 'PS-N:' prefix directly.
    """
    out: list[Zone] = []
    # Build {zone_num: rsp_num} lookup from RSP zone ranges
    zone_to_rsp: dict[int, int] = {}
    for r in rsps:
        for z in r.zones:
            zone_to_rsp[z] = r.number

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row or not row[0]:
            continue
        z_label = str(row[0]).strip()
        if not (z_label.startswith("Z") and z_label[1:].isdigit()):
            continue
        zone_num = int(z_label[1:])

        desc = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if not desc:
            continue

        is_spare = desc.upper() == "SPARE"
        m_ac = _PS_AC_RE.match(desc)
        m_batt = _PS_BATT_RE.match(desc)
        is_ps_ac = bool(m_ac)
        is_ps_batt = bool(m_batt)

        if is_ps_ac:
            rsp_number = int(m_ac.group(1))
        elif is_ps_batt:
            rsp_number = int(m_batt.group(1))
        elif is_spare:
            rsp_number = None
        else:
            rsp_number = zone_to_rsp.get(zone_num)

        out.append(Zone(
            number=zone_num,
            description=desc,
            rsp_number=rsp_number,
            is_spare=is_spare,
            is_ps_ac=is_ps_ac,
            is_ps_batt=is_ps_batt,
        ))
    return out


def _master_zones_from_point_info(zones: list[ZoneInfo], rsps: list[RSP]) -> list[Zone]:
    """Reconstruct Master-sheet zone rows from the per-RSP Point Info sheets.

    Fallback for older worksheets that have no 'Master' sheet. parse_dmp_worksheet
    already reads the 'DMP 714-16 Point Info' sheets into ZoneInfo (number, room
    location, device_type); this turns them into the Zone list the door chart
    consumes, mirroring what generate_dmp_ws.py writes into Master so that
    inject() and the door chart's conditional formatting behave identically.

    The exact phrases 'A/C LOSS' / 'BATT. TRBL' (and the 'PS-N:' prefix) are
    required: the door chart's conditional formatting keys on them, and a later
    re-parse of a regenerated Master via _PS_AC_RE/_PS_BATT_RE expects that shape.
    """
    zone_to_rsp: dict[int, int] = {z: r.number for r in rsps for z in r.zones}

    out: list[Zone] = []
    for zi in sorted(zones, key=lambda z: z.number):
        loc = (zi.location or "").strip()
        loc_up = loc.upper()
        dtype = (zi.device_type or "").strip().lower()
        rsp_num = zone_to_rsp.get(zi.number)

        if loc_up == "SPARE":
            out.append(Zone(
                number=zi.number, description="SPARE",
                rsp_number=None, is_spare=True,
            ))
        elif dtype == "supervisory" or "TROUBLE" in loc_up:
            ps = f"PS-{rsp_num}" if rsp_num else "PS"
            if "BATT" in loc_up:
                out.append(Zone(
                    number=zi.number, description=f"{ps}: BATT. TRBL",
                    rsp_number=rsp_num, is_ps_batt=True,
                ))
            else:
                out.append(Zone(
                    number=zi.number, description=f"{ps}: A/C LOSS",
                    rsp_number=rsp_num, is_ps_ac=True,
                ))
        else:
            out.append(Zone(
                number=zi.number, description=loc,
                rsp_number=rsp_num,
            ))
    return out


# -------- CLI for quick testing --------

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("usage: python parse_dmp_worksheet.py <dmp_worksheet.xlsx>")
        sys.exit(1)

    design = parse_dmp_worksheet(sys.argv[1])

    print("=== Site Info ===")
    print(f"  School: {design.site_info.school_name}")
    print(f"  Code: {design.site_info.school_code}")
    print(f"  Phone: {design.site_info.phone}")
    print(f"  Install Tech: {design.site_info.install_tech}")
    print(f"  XR550 Location: {design.site_info.xr550_location}")

    print(f"\n=== Splitters ({len(design.splitters)}) ===")
    for s in design.splitters:
        print(f"  {s.id} ({s.splitter_type}) @ {s.location}")
        for inp_name, inp_desc in s.inputs.items():
            print(f"    IN: {inp_name} = {inp_desc}")
        for out_desc in s.outputs:
            print(f"    OUT: {out_desc}")

    print(f"\n=== Keypads ({len(design.keypads)}) ===")
    for k in design.keypads:
        print(f"  KP{k.number}: source={k.source}, location={k.location}, global={k.global_keypad}")

    print(f"\n=== Power Supplies ({len(design.power_supplies)}) ===")
    for ps in design.power_supplies:
        print(f"  PS{ps.number} @ {ps.location}")
        for relay_num, relay_func in sorted(ps.relays.items()):
            print(f"    RELAY {relay_num}: {relay_func}")

    print(f"\n=== Zones ({len(design.zones)}) ===")
    for z in design.zones[:20]:
        print(f"  Z{z.number}: {z.location} ({z.device_type})")
    if len(design.zones) > 20:
        print(f"  ... ({len(design.zones) - 20} more)")

    print(f"\n=== Master Zones ({len(design.master_zones)}) ===")
    for z in design.master_zones[:15]:
        flags = []
        if z.is_spare: flags.append("SPARE")
        if z.is_ps_ac: flags.append("PS_AC")
        if z.is_ps_batt: flags.append("PS_BATT")
        flag_str = f" [{','.join(flags)}]" if flags else ""
        print(f"  Z{z.number}: {z.description!r} rsp={z.rsp_number}{flag_str}")
    if len(design.master_zones) > 15:
        print(f"  ... ({len(design.master_zones) - 15} more)")
