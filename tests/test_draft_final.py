"""Tests for DRAFT/FINAL stamping of generated DMP worksheets.

The stamp is the lifecycle boundary: DRAFT exports carry a visible
not-for-install banner plus a DMPStatus doc property the import path refuses;
FINAL carries only the doc property. Both must survive write_dmp_xlsx's
template-overlay post-processing.

Run: pytest tests/test_draft_final.py
"""
from datetime import date
from pathlib import Path
import sys

import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent
SCRIPTS = REPO_ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))

from generate_dmp_ws import dmp_filename, write_dmp_xlsx  # noqa: E402
from parse_dmp_worksheet import (  # noqa: E402
    DMPDesign,
    RSP,
    SiteInfo,
    ZoneInfo,
    parse_dmp_worksheet,
)

DMP_TEMPLATE = REPO_ROOT / "DMP Installation Worksheet_template_blank.xlsx"

pytestmark = pytest.mark.skipif(
    not DMP_TEMPLATE.exists(),
    reason="DMP worksheet template fixture not present",
)


def _design() -> DMPDesign:
    return DMPDesign(
        site_info=SiteInfo(school_name="STAMP TEST SCHOOL",
                           ip_address="10.0.0.2", default_gateway="10.0.0.1"),
        rsps=[RSP(number=1, location="FACP ROOM", zones=[501, 502])],
        zones=[
            ZoneInfo(number=501, location="FACP ROOM", device_type="Motion", partition=1),
            ZoneInfo(number=502, location="SPARE", device_type="Spare", partition=1),
        ],
    )


# -------- filename helper --------

def test_dmp_filename_stamps():
    assert dmp_filename("X", date_str="2026-06-10") == "X_dmp_2026-06-10.xlsx"
    assert dmp_filename("X", "DRAFT", "2026-06-10") == "X_dmp_DRAFT_2026-06-10.xlsx"
    assert dmp_filename("X", "FINAL", "2026-06-10") == "X_dmp_FINAL_2026-06-10.xlsx"


def test_dmp_filename_defaults_to_today():
    assert date.today().isoformat() in dmp_filename("X", "DRAFT")


# -------- stamped output round-trips --------

def test_draft_output_carries_status_and_banner(tmp_path):
    out = tmp_path / "draft.xlsx"
    write_dmp_xlsx(_design(), DMP_TEMPLATE, out, stamp="DRAFT")

    design = parse_dmp_worksheet(out)
    assert design.dmp_status == "DRAFT"

    import openpyxl
    ws = openpyxl.load_workbook(out)["SITE INFO"]
    assert "DRAFT" in (ws["A1"].value or "")
    assert "NOT FOR INSTALL" in ws["A1"].value


def test_final_output_carries_status_no_banner(tmp_path):
    out = tmp_path / "final.xlsx"
    write_dmp_xlsx(_design(), DMP_TEMPLATE, out, stamp="FINAL")

    design = parse_dmp_worksheet(out)
    assert design.dmp_status == "FINAL"

    import openpyxl
    ws = openpyxl.load_workbook(out)["SITE INFO"]
    assert ws["A1"].value is None


def test_unstamped_output_has_no_status(tmp_path):
    out = tmp_path / "plain.xlsx"
    write_dmp_xlsx(_design(), DMP_TEMPLATE, out)
    design = parse_dmp_worksheet(out)
    assert design.dmp_status == ""
