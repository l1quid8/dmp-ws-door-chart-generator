"""Every splitter in the DMP reaches the door chart, whatever its number.

hardware.next_splitter_id names every LX splitter '710-LX500-N' for N up to
MAX_SPLITTERS_PER_TYPE, but the door chart template's pre-seeded slot IDs spread
the LX rows across five buses (LX500-1..5, LX600-1..5, ... LX900-1..5). Population
used to walk those slots and match by exact ID, so a sixth LX500 splitter matched
nothing and vanished from the deliverable with no error — the HAYNES_CHARTER_ES
job silently lost 710-LX500-6 and 710-LX500-7 (and RSP-5/RSP-6's feed with them).

Run: pytest tests/test_splitter_overflow.py
"""
from pathlib import Path
import sys

import openpyxl
import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent
SCRIPTS = REPO_ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))

from hardware import MAX_SPLITTERS_PER_TYPE  # noqa: E402
from inject_door_chart import inject  # noqa: E402
from test_door_chart_consolidation import _kp, _lx, _splitter_design  # noqa: E402

DOOR_CHART_TEMPLATE = REPO_ROOT / "door_chart_template_blank.xlsx"

pytestmark = pytest.mark.skipif(not DOOR_CHART_TEMPLATE.exists(),
                                reason="door chart template fixture not present")

TOPOLOGY_ROWS = range(29, 64)


def _topology(out_path):
    """{splitter id: (location, combus input, out1, out2, out3)} from Master 29-63."""
    m = openpyxl.load_workbook(out_path)["Master"]
    found = {}
    for r in TOPOLOGY_ROWS:
        sid = m[f"A{r}"].value
        if sid and m[f"C{r}"].value:      # col C (section title) only set when populated
            found[str(sid).strip()] = tuple(
                m[f"{c}{r}"].value for c in ("B", "D", "E", "F", "G"))
    return found


def _inject(tmp_path, design):
    out = tmp_path / "door_chart.xlsx"
    inject(DOOR_CHART_TEMPLATE, design, out)
    return out


def _haynes():
    """The real HAYNES_CHARTER_ES topology: 7 LX500 splitters and 3 KP."""
    return _splitter_design(
        _lx(1, "ADMIN BUILDING MAIN OFFICE"),
        _lx(2, "ADMIN BUILDING MAIN OFFICE"),
        _lx(3, "MULTIPURPOSE BUILDING (PASSAGE)"),
        _lx(4, "BUILDING C CLASSROOM 16"),
        _lx(5, "BUILDING A ELEC. SWITCH ROOM"),
        _lx(6, "BUILDING B CUSTODIAN"),
        _lx(7, "KINDERGARTEN BUILDING 1 CLASSROOM 3"),
        _kp(1), _kp(2), _kp(3),
    )


def test_sixth_and_seventh_lx500_reach_the_chart(tmp_path):
    """The reported bug: LX500-6/-7 dropped out of the door chart entirely."""
    found = _topology(_inject(tmp_path, _haynes()))
    assert "710-LX500-6" in found, "710-LX500-6 missing from door chart topology"
    assert "710-LX500-7" in found, "710-LX500-7 missing from door chart topology"
    assert found["710-LX500-6"][0] == "BUILDING B CUSTODIAN"
    assert found["710-LX500-7"][0] == "KINDERGARTEN BUILDING 1 CLASSROOM 3"


def test_all_haynes_splitters_present_and_ordered(tmp_path):
    out = _inject(tmp_path, _haynes())
    m = openpyxl.load_workbook(out)["Master"]
    ids = [str(m[f"A{r}"].value).strip() for r in TOPOLOGY_ROWS
           if m[f"A{r}"].value and m[f"C{r}"].value]
    assert ids == [f"710-LX500-{n}" for n in range(1, 8)] + \
                  [f"710-KP-{n}" for n in range(1, 4)]


def test_chart_tab_renders_every_splitter(tmp_path):
    """Master rows are useless if the LX-KP-710s tab doesn't point at them."""
    out = _inject(tmp_path, _haynes())
    wb = openpyxl.load_workbook(out)
    m, lx = wb["Master"], wb["LX-KP-710s"]
    charted = {v for row in lx.iter_rows() for c in row
               if isinstance(v := c.value, str) and v.startswith("=Master!C")}
    row_to_id = {f"=Master!C{r}": str(m[f'A{r}'].value).strip() for r in TOPOLOGY_ROWS}
    assert {row_to_id[c] for c in charted} == {
        *(f"710-LX500-{n}" for n in range(1, 8)),
        *(f"710-KP-{n}" for n in range(1, 4)),
    }


def test_max_splitters_per_type_all_fit(tmp_path):
    """hardware.py lets the user create this many; none may be silently dropped."""
    n = MAX_SPLITTERS_PER_TYPE
    design = _splitter_design(*(_lx(i) for i in range(1, n + 1)),
                              *(_kp(i) for i in range(1, n + 1)))
    found = _topology(_inject(tmp_path, design))
    assert len(found) == 2 * n
    for i in range(1, n + 1):
        assert f"710-LX500-{i}" in found
        assert f"710-KP-{i}" in found


def test_multi_bus_design_still_works(tmp_path):
    """Designs that genuinely use LX600/LX700 keep working, grouped by bus."""
    design = _splitter_design(_lx(1), _lx(2), _lx(1, bus=600), _lx(1, bus=700), _kp(1))
    found = _topology(_inject(tmp_path, design))
    assert set(found) == {"710-LX500-1", "710-LX500-2", "710-LX600-1",
                          "710-LX700-1", "710-KP-1"}
