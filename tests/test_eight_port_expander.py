"""Regression tests for 8-port (714-8) zone expanders.

A 714-8 module owns only 8 contiguous zones, but the worksheet template's
per-module Point Info sheets are hard-wired to a fixed 16-zone stride, and the
PDF-parse path hardcoded supervisory-zone offsets (+14/+15) and the "714-16"
model. With packed/sequential zone numbering (the CAD convention), an 8-port
module before the last module shifted every later Point Info sheet off its
zones and mislabeled power-supply relays.

These tests pin the fix: Point Info sheets map to each module's *actual* zones,
8-port modules leave their unused rows blank, relay labels reference the real
supervisory zones, and the model reads "714-8".

Run: pytest tests/test_eight_port_expander.py
"""
from pathlib import Path
import sys

import openpyxl
import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent
SCRIPTS = REPO_ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))

from parse_dmp_worksheet import DMPDesign, RSP, PowerSupply, ZoneInfo  # noqa: E402
from generate_dmp_ws import (  # noqa: E402
    write_dmp_xlsx,
    _expander_model_for_count,
    _ps_relays,
)

TEMPLATE = REPO_ROOT / "DMP Installation Worksheet_template_blank.xlsx"

pytestmark = pytest.mark.skipif(
    not TEMPLATE.exists(), reason="worksheet template fixture not present"
)


# -------- pure helpers (no file dependency) --------

def test_model_inferred_from_zone_count():
    assert _expander_model_for_count(8) == "714-8"
    assert _expander_model_for_count(16) == "714-16"
    # fewer than a full 8-port (rare/partial) still reads as 714-8
    assert _expander_model_for_count(6) == "714-8"


def test_ps_relays_use_actual_supervisory_zones():
    # 8-port module at 517-524: supervisory zones are 523/524, NOT 531/532.
    relays = _ps_relays(list(range(517, 525)), "714-8", 2)
    assert "523" in relays[2] and "A/C" in relays[2].upper().replace("AC", "A/C")
    assert "524" in relays[3]
    assert "714-8" in relays[2] and "714-8" in relays[3]
    assert "531" not in relays[2] and "532" not in relays[3]

    # 16-port module keeps its 515/516 supervisory pair and 714-16 label.
    relays16 = _ps_relays(list(range(501, 517)), "714-16", 1)
    assert "515" in relays16[2] and "516" in relays16[3]
    assert "714-16" in relays16[2]


# -------- Point Info dynamic mapping (end-to-end via write_dmp_xlsx) --------

def _mixed_design():
    """RSP1=714-16 (501-516), RSP2=714-8 (517-524), RSP3=714-16 (525-540) — packed."""
    design = DMPDesign()
    design.site_info.school_name = "TEST SCHOOL"
    specs = [
        (1, "714-16", range(501, 517)),
        (2, "714-8", range(517, 525)),
        (3, "714-16", range(525, 541)),
    ]
    for num, model, zrange in specs:
        zones = list(zrange)
        design.rsps.append(RSP(number=num, location=f"RSP {num}", zones=zones, model=model))
        design.power_supplies.append(
            PowerSupply(number=num, location=f"RSP {num}", relays=_ps_relays(zones, model, num))
        )
        for i, zn in enumerate(zones):
            if i >= len(zones) - 2:
                dt, loc = "Supervisory", ("PS A/C" if i == len(zones) - 2 else "PS BATT")
            else:
                dt, loc = "Motion", f"ROOM {zn}"
            design.zones.append(ZoneInfo(number=zn, location=loc, device_type=dt, partition=1))
    return design


def _formula_target_row(value):
    """'=Master!A18' -> 18 ; None/blank -> None."""
    if not isinstance(value, str) or "Master!" not in value:
        return None
    return int(value.split("!A")[1])


def test_point_info_maps_each_module_to_its_real_zones(tmp_path):
    out = tmp_path / "ws.xlsx"
    write_dmp_xlsx(_mixed_design(), TEMPLATE, out)
    wb = openpyxl.load_workbook(out, data_only=False)

    # Master row of a zone: Z501 -> row 2, so Zn -> row (n - 499).
    def mrow(n):
        return n - 499

    pi2 = wb["DMP 714-16 Point Info (2)"]
    # Row 4 = first zone (517); 8-port fills rows 4-11 only.
    assert _formula_target_row(pi2["A4"].value) == mrow(517)
    assert _formula_target_row(pi2["A11"].value) == mrow(524)  # last real zone
    assert pi2["F10"].value == "Supervisory" and pi2["F11"].value == "Supervisory"
    # Rows 12-19 must be blank — the bug pulled RSP3's zones (525+) in here.
    for r in range(12, 20):
        assert pi2[f"A{r}"].value in (None, "")
        assert pi2[f"B{r}"].value in (None, "")

    pi3 = wb["DMP 714-16 Point Info (3)"]
    # RSP3 starts at 525 (packed), NOT 533 (the old fixed stride).
    assert _formula_target_row(pi3["A4"].value) == mrow(525)
    assert _formula_target_row(pi3["A19"].value) == mrow(540)  # 16 zones fill 4-19

    # Source cell reflects the model.
    assert "714-8" in str(pi2["A20"].value)


def test_all_sixteen_port_design_keeps_original_stride(tmp_path):
    """No regression: when every module is 714-16, Point Info keeps the 16-stride."""
    design = DMPDesign()
    for num, start in [(1, 501), (2, 517), (3, 533)]:
        zones = list(range(start, start + 16))
        design.rsps.append(RSP(number=num, location=f"RSP {num}", zones=zones, model="714-16"))
        for i, zn in enumerate(zones):
            dt = "Supervisory" if i >= 14 else "Motion"
            design.zones.append(ZoneInfo(number=zn, location=f"ROOM {zn}", device_type=dt, partition=1))
    out = tmp_path / "ws16.xlsx"
    write_dmp_xlsx(design, TEMPLATE, out)
    wb = openpyxl.load_workbook(out, data_only=False)
    assert _formula_target_row(wb["DMP 714-16 Point Info (1)"]["A4"].value) == 2   # Z501
    assert _formula_target_row(wb["DMP 714-16 Point Info (2)"]["A4"].value) == 18  # Z517
    assert _formula_target_row(wb["DMP 714-16 Point Info (3)"]["A4"].value) == 34  # Z533


def test_export_normalizes_stale_model_from_zone_count(tmp_path):
    """A worksheet re-imported from a pre-fix file carries a stale model="714-16" on its
    8-port modules (read from the old Exp Mod text). Export must re-derive the model and
    relay text from the live zone count so every sheet reads "714-8"."""
    design = DMPDesign()
    # RSP 2 is physically an 8-port (8 zones) but mislabeled 714-16, with stale relay text.
    zones = list(range(517, 525))
    design.rsps.append(
        RSP(number=2, location="RSP 2", zones=zones, model="714-16")  # stale
    )
    design.power_supplies.append(
        PowerSupply(number=2, location="RSP 2", relays={
            1: "12v DC Output to Terminal Strip",
            2: "AC Trouble Zone 531 (714-16 Expander #2)",   # stale, wrong zone
            3: "Battery Trouble Zone 532 (714-16 Expander #2)",
            4: "Battery 12V",
        })
    )
    for i, zn in enumerate(zones):
        dt = "Supervisory" if i >= len(zones) - 2 else "Motion"
        design.zones.append(ZoneInfo(number=zn, location=f"ROOM {zn}", device_type=dt, partition=1))

    out = tmp_path / "stale.xlsx"
    write_dmp_xlsx(design, TEMPLATE, out)
    wb = openpyxl.load_workbook(out, data_only=False)

    # Point Info row 20 (the user's report)
    assert "714-8" in str(wb["DMP 714-16 Point Info (2)"]["A20"].value)
    # Exp Mod prefix
    em = wb["DMP 714 Exp Mod"]
    prefix = next(em[f"A{r}"].value for r in range(4, 12) if em[f"B{r}"].value == 2)
    assert prefix == "DMP 714-8 #"
    # Power Supply relays — correct supervisory zones + model, not the stale 531/532/714-16
    ps = wb["DMP 505-12_G Power Supply 1-10"]
    relay_text = " ".join(str(ps[f"C{r}"].value) for r in range(2, 8))
    assert "523" in relay_text and "524" in relay_text and "714-8" in relay_text
    assert "531" not in relay_text and "714-16" not in relay_text
