"""Door chart consolidation: empty placeholder blocks are removed, not just blanked.

The template pre-draws 15 block groups per presentation sheet. inject() truncates each
sheet right after its last filled group, drops the orphaned Excel Tables and logo
drawing anchors, and compacts the LX-KP-710s tab's scattered splitter charts into the
earliest slots — matching the hand-finished deliverables.

Run: pytest tests/test_door_chart_consolidation.py
"""
from pathlib import Path
import posixpath
import re
import sys
import zipfile

import openpyxl
import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent
SCRIPTS = REPO_ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))

from parse_dmp_worksheet import DMPDesign, Splitter  # noqa: E402
from inject_door_chart import inject  # noqa: E402
from test_door_chart_eight_port import _mixed_design  # noqa: E402

DOOR_CHART_TEMPLATE = REPO_ROOT / "door_chart_template_blank.xlsx"

pytestmark = pytest.mark.skipif(not DOOR_CHART_TEMPLATE.exists(),
                                reason="door chart template fixture not present")

SHEET_TABS = {3: "Terminal Cans", 4: "RSPs", 5: "Power Supplies", 6: "LX-KP-710s"}


def _lx(num, loc="ADMIN BUILDING", bus=500):
    return Splitter(id=f"710-LX{bus}-{num}", splitter_type="LX", location=loc)


def _kp(num, loc="ADMIN BUILDING"):
    return Splitter(id=f"710-KP-{num}", splitter_type="KP", location=loc)


def _splitter_design(*splitters):
    design = _mixed_design()
    design.splitters = list(splitters)
    return design


def _inject(tmp_path, design):
    out = tmp_path / "door_chart.xlsx"
    inject(DOOR_CHART_TEMPLATE, design, out)
    return out


def _max_sheet_row(zf, sheet_num):
    xml = zf.read(f"xl/worksheets/sheet{sheet_num}.xml").decode()
    return max(int(m.group(1)) for m in re.finditer(r'<row r="(\d+)"', xml))


# -------- template + MSP --------

def test_msp_tab_absent_and_master_intact(tmp_path):
    assert "MSP" not in openpyxl.load_workbook(DOOR_CHART_TEMPLATE).sheetnames
    out = _inject(tmp_path, _mixed_design())
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Header", "Master", "Terminal Cans", "RSPs",
                             "Power Supplies", "LX-KP-710s"]
    m = wb["Master"]
    assert m["A2"].value == "XR-550 CONFIGURATION"       # config sub-table survives
    assert m["A27"].value == "710 BUS SPLITTER/REPEATER TOPOLOGY"


# -------- truncation --------

def test_truncation_bounds_mixed_design(tmp_path):
    """3 RSPs (max module 3) and no splitters -> 2 groups on TC/RSPs/PS, 1 on LX."""
    out = _inject(tmp_path, _mixed_design())
    zf = zipfile.ZipFile(out)
    assert _max_sheet_row(zf, 3) == 51   # Terminal Cans: group 2 ends at 27+24
    assert _max_sheet_row(zf, 4) == 51   # RSPs: 28+23
    assert _max_sheet_row(zf, 5) == 24   # Power Supplies: 14+10
    assert _max_sheet_row(zf, 6) == 12   # LX: no splitters, one blanked group kept

    wb = openpyxl.load_workbook(out)
    for sheet_num, tab in SHEET_TABS.items():
        ws = wb[tab]
        assert ws.max_row == _max_sheet_row(zf, sheet_num)
        # nothing below the cutoff — no block scaffolding survives
        hdrs = [c.row for row in ws.iter_rows() for c in row if c.value == "=Header!B3"]
        assert hdrs and max(hdrs) <= ws.max_row


def test_no_formulas_or_headers_below_cutoff(tmp_path):
    out = _inject(tmp_path, _mixed_design())
    zf = zipfile.ZipFile(out)
    for sheet_num, cutoff in ((3, 51), (4, 51), (5, 24), (6, 12)):
        xml = zf.read(f"xl/worksheets/sheet{sheet_num}.xml").decode()
        for m in re.finditer(r'<c r="[A-Z]+(\d+)"[^>]*?><f>(Master|Header)!', xml):
            assert int(m.group(1)) <= cutoff, \
                f"sheet{sheet_num}: formula cell on row {m.group(1)} survived truncation"


# -------- LX-KP-710s compaction --------

def test_lx_compaction_pairs_scattered_splitters(tmp_path):
    """LX500-1 and KP-1 are paired side by side in group 1 (the hand-finished layout).

    Master packs splitters contiguously from row 29 in display order, so KP-1 backs
    row 30 — see tests/test_splitter_overflow.py for why it is no longer row 54."""
    out = _inject(tmp_path, _splitter_design(_lx(1), _kp(1)))
    wb = openpyxl.load_workbook(out)
    lx, m = wb["LX-KP-710s"], wb["Master"]
    assert (m["A29"].value, m["A30"].value) == ("710-LX500-1", "710-KP-1")
    assert lx["B7"].value == "=Master!C29"
    assert lx["E7"].value == "=Master!C30"
    assert [lx[f"F{r}"].value for r in range(9, 13)] == \
        ["=Master!D30", "=Master!E30", "=Master!F30", "=Master!G30"]
    assert lx.max_row == 12


def test_lx_odd_count_blanks_right_slot_keeps_table_header(tmp_path):
    out = _inject(tmp_path, _splitter_design(_lx(1)))
    lx = openpyxl.load_workbook(out)["LX-KP-710s"]
    assert lx["B7"].value == "=Master!C29"
    assert lx["E7"].value is None                      # unused right slot blanked
    assert all(lx[f"F{r}"].value is None for r in range(9, 13))
    assert lx["E8"].value and lx["F8"].value           # Excel Table header row kept
    assert lx.max_row == 12


def test_lx_three_splitters_spill_to_group_two(tmp_path):
    out = _inject(tmp_path, _splitter_design(_lx(1), _lx(2), _kp(1)))
    lx = openpyxl.load_workbook(out)["LX-KP-710s"]
    assert lx["B7"].value == "=Master!C29"
    assert lx["E7"].value == "=Master!C30"
    assert lx["B19"].value == "=Master!C31"            # 3rd chart: group 2 left slot
    assert lx["E19"].value is None
    assert lx.max_row == 24


# -------- package consistency (the anti-"repair file" test) --------

def _rel_targets(zf, rels_name):
    base = posixpath.dirname(posixpath.dirname(rels_name))  # strip _rels/
    for m in re.finditer(r'<Relationship\b[^>]*?/>', zf.read(rels_name).decode()):
        target = re.search(r'Target="([^"]+)"', m.group(0)).group(1)
        mode = re.search(r'TargetMode="External"', m.group(0))
        if not mode:
            yield posixpath.normpath(posixpath.join(base, target))


def test_package_consistency(tmp_path):
    out = _inject(tmp_path, _splitter_design(_lx(1), _kp(1)))
    zf = zipfile.ZipFile(out)
    names = set(zf.namelist())

    assert "xl/calcChain.xml" not in names

    # Every Content-Types Override points at an existing part.
    ct = zf.read("[Content_Types].xml").decode()
    for m in re.finditer(r'PartName="/([^"]+)"', ct):
        assert m.group(1) in names, f"Content-Types Override for missing part {m.group(1)}"

    # Every internal Relationship target exists; collect referenced tables.
    referenced = set()
    for rels_name in [n for n in names if n.endswith(".rels")]:
        for target in _rel_targets(zf, rels_name):
            assert target in names, f"{rels_name} points at missing part {target}"
            if target.startswith("xl/tables/"):
                referenced.add(target)

    # No orphan table parts, and each surviving table has a Content-Types Override.
    tables = {n for n in names if n.startswith("xl/tables/")}
    assert tables == referenced
    for t in tables:
        assert f'PartName="/{t}"' in ct

    for sheet_num in SHEET_TABS:
        sheet = f"xl/worksheets/sheet{sheet_num}.xml"
        xml = zf.read(sheet).decode()
        rels = zf.read(f"xl/worksheets/_rels/sheet{sheet_num}.xml.rels").decode()
        rel_ids = set(re.findall(r'Id="(rId\d+)"', rels))
        dim_end = int(re.search(r'<dimension ref="[A-Z]+\d+:[A-Z]+(\d+)"/>', xml).group(1))

        # every tablePart rId resolves; count attribute is accurate
        tp_ids = re.findall(r'<tablePart r:id="(rId\d+)"/>', xml)
        assert set(tp_ids) <= rel_ids
        tp_count = re.search(r'<tableParts count="(\d+)">', xml)
        assert tp_count and int(tp_count.group(1)) == len(tp_ids)

        # rows, merges and surviving-table refs all lie within the dimension
        assert max(int(r) for r in re.findall(r'<row r="(\d+)"', xml)) <= dim_end
        mc = re.search(r'<mergeCells count="(\d+)">(.*?)</mergeCells>', xml, re.S)
        merges = re.findall(r'<mergeCell ref="[A-Z]+\d+:[A-Z]+(\d+)"/>', mc.group(2))
        assert int(mc.group(1)) == len(merges)
        assert all(int(r) <= dim_end for r in merges)
        for target in _rel_targets(zf, f"xl/worksheets/_rels/sheet{sheet_num}.xml.rels"):
            if target.startswith("xl/tables/"):
                ref_end = int(re.search(r'<table [^>]*?\bref="[A-Z]+\d+:[A-Z]+(\d+)"',
                                        zf.read(target).decode()).group(1))
                assert ref_end <= dim_end, f"{target} extends past sheet{sheet_num} dimension"

        # drawing anchors only on kept rows (from-row is 0-based)
        drawing = next(t for t in _rel_targets(zf, f"xl/worksheets/_rels/sheet{sheet_num}.xml.rels")
                       if t.startswith("xl/drawings/"))
        for m in re.finditer(r"<xdr:from>.*?<xdr:row>(\d+)</xdr:row>", zf.read(drawing).decode(), re.S):
            assert int(m.group(1)) < dim_end


def test_openpyxl_roundtrip_tables_parse(tmp_path):
    out = _inject(tmp_path, _splitter_design(_lx(1), _kp(1)))
    wb = openpyxl.load_workbook(out)   # validates zip + table refs against sheet bounds
    for tab in SHEET_TABS.values():
        assert wb[tab].tables            # surviving tables parse and are non-empty
