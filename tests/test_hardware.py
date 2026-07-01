"""Tests for hardware.py — post-CAD add/remove of expanders/splitters/keypads.

The zone-block contract is the load-bearing part: expander module N always
owns the fixed 16-zone address stride (Point Info sheet N is hard-wired to
it), a 714-8 materializes only 8 points, and removal leaves numbering gaps
rather than renumbering physical addresses.

Run: pytest tests/test_hardware.py
"""
from pathlib import Path
import sys

import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent
SCRIPTS = REPO_ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))

from parse_dmp_worksheet import (  # noqa: E402
    DMPDesign,
    Keypad,
    PowerSupply,
    RSP,
    SiteInfo,
    Splitter,
    ZoneInfo,
)
from hardware import (  # noqa: E402
    HardwareError,
    MAX_EXPANDERS,
    MAX_KEYPADS,
    MAX_SPLITTERS_PER_TYPE,
    add_expander,
    add_keypad,
    add_splitter,
    block_orphans,
    diff_refs,
    existing_locations,
    snapshot_refs,
    next_expander_number,
    remove_expander,
    remove_keypad,
    remove_splitter,
    renumber_splitter,
    zone_block_for,
)

DMP_TEMPLATE = REPO_ROOT / "DMP Installation Worksheet_template_blank.xlsx"


def _design_with_expanders(n: int) -> DMPDesign:
    d = DMPDesign(site_info=SiteInfo(school_name="HW TEST"))
    for _ in range(n):
        add_expander(d, "714-16")
    return d


# -------- add_expander --------

def test_add_714_16():
    d = _design_with_expanders(0)
    rsp = add_expander(d, "714-16", location="FACP ROOM")
    assert rsp.number == 1 and rsp.model == "714-16"
    assert rsp.zones == list(range(501, 517))
    assert d.power_supplies[0].number == 1
    by_num = {z.number: z for z in d.zones}
    assert len(d.zones) == 16
    assert by_num[501].location == "SPARE" and by_num[501].device_type == "Spare"
    assert by_num[515].location == "PS-1: A/C LOSS"
    assert by_num[516].location == "PS-1: BATT. TRBL"
    assert by_num[515].device_type == "Supervisory"


def test_add_714_8_consumes_full_address_block():
    d = _design_with_expanders(1)
    rsp = add_expander(d, "714-8")
    assert rsp.number == 2 and rsp.model == "714-8"
    # 8 real points inside module 2's 16-zone stride (Z517..Z532)
    assert rsp.zones == list(range(517, 525))
    by_num = {z.number: z for z in d.zones}
    assert by_num[523].location == "PS-2: A/C LOSS"
    assert by_num[524].location == "PS-2: BATT. TRBL"
    assert 525 not in by_num            # rest of the block unallocated
    # A later expander still gets the NEXT stride, not a packed one
    rsp3 = add_expander(d, "714-16")
    assert rsp3.zones[0] == 533


def test_unknown_model_rejected():
    with pytest.raises(HardwareError):
        add_expander(DMPDesign(), "714-32")


def test_expander_capacity_guard():
    d = _design_with_expanders(MAX_EXPANDERS)
    with pytest.raises(HardwareError, match="15"):
        add_expander(d, "714-16")


def test_gap_reuse_after_removal():
    d = _design_with_expanders(3)
    remove_expander(d, 2)
    assert next_expander_number(d) == 2
    rsp = add_expander(d, "714-16")
    assert rsp.number == 2
    assert rsp.zones[0] == 517          # the freed Z517 block


def test_bus_boundary_addressing():
    """Bus 500 carries modules 1-6 (Z501-596); module 7 starts bus 600 at
    Z601 (the template's Master skips Z597-600), module 13 starts bus 700."""
    assert list(zone_block_for(6))[0] == 581
    assert list(zone_block_for(7))[0] == 601
    assert list(zone_block_for(12))[0] == 681
    assert list(zone_block_for(13))[0] == 701
    assert list(zone_block_for(15)) == list(range(733, 749))  # matches ACADEMY

    d = _design_with_expanders(6)
    rsp7 = add_expander(d, "714-8")
    assert rsp7.zones == list(range(601, 609))
    by_num = {z.number: z for z in d.zones}
    assert by_num[607].location == "PS-7: A/C LOSS"
    assert by_num[608].location == "PS-7: BATT. TRBL"


# -------- remove_expander --------

def test_remove_expander_drops_block_and_scrubs_outputs():
    d = _design_with_expanders(2)
    d.splitters.append(Splitter(id="710-LX500-1", splitter_type="LX",
                                outputs=["RSP-1", "RSP-2", "Spare"]))
    remove_expander(d, 2)
    assert [r.number for r in d.rsps] == [1]
    assert [p.number for p in d.power_supplies] == [1]
    assert max(z.number for z in d.zones) == 516
    assert d.splitters[0].outputs == ["RSP-1", "Spare", "Spare"]


def test_remove_missing_expander_raises():
    with pytest.raises(HardwareError):
        remove_expander(_design_with_expanders(1), 9)


# -------- splitters --------

def test_add_splitter_ids_and_capacity():
    d = DMPDesign()
    s1 = add_splitter(d, "LX", location="FACP")
    s2 = add_splitter(d, "LX")
    k1 = add_splitter(d, "KP")
    assert s1.id == "710-LX500-1" and s2.id == "710-LX500-2"
    assert k1.id == "710-KP-1"
    assert s1.outputs == ["Spare", "Spare", "Spare"]
    for _ in range(MAX_SPLITTERS_PER_TYPE - 2):
        add_splitter(d, "LX")
    with pytest.raises(HardwareError):
        add_splitter(d, "LX")
    add_splitter(d, "KP")               # KP pool is separate


def test_splitter_id_gap_reuse():
    d = DMPDesign()
    add_splitter(d, "LX"); add_splitter(d, "LX"); add_splitter(d, "LX")
    remove_splitter(d, "710-LX500-2")
    assert add_splitter(d, "LX").id == "710-LX500-2"


def test_remove_splitter_scrubs_refs_and_keypad_sources():
    d = DMPDesign()
    s1 = add_splitter(d, "LX")
    s2 = add_splitter(d, "LX")
    s1.outputs = ["To 710-LX500-2", "RSP-1", "Spare"]
    d.keypads.append(Keypad(number=2, source="710-LX500-2", location="HALL"))
    remove_splitter(d, "710-LX500-2")
    assert s1.outputs == ["Spare", "RSP-1", "Spare"]
    assert d.keypads[0].source is None


def test_renumber_splitter_into_free_slot():
    d = DMPDesign()
    add_splitter(d, "LX")
    renumber_splitter(d, "710-LX500-1", 4)
    assert [s.id for s in d.splitters] == ["710-LX500-4"]


def test_renumber_splitter_rewrites_all_refs():
    d = DMPDesign()
    s1 = add_splitter(d, "LX")          # 710-LX500-1
    s2 = add_splitter(d, "LX")          # 710-LX500-2
    s1.outputs = ["To 710-LX500-2", "RSP-1", "Spare"]
    s2.inputs = {"LX-Bus In": "From 710-LX500-1"}
    d.keypads.append(Keypad(number=2, source="710-LX500-2", location="HALL"))
    renumber_splitter(d, "710-LX500-2", 5)
    assert s1.outputs == ["To 710-LX500-5", "RSP-1", "Spare"]
    assert s2.id == "710-LX500-5"
    assert d.keypads[0].source == "710-LX500-5"


def test_renumber_splitter_rejects_duplicate():
    d = DMPDesign()
    add_splitter(d, "LX"); add_splitter(d, "LX")
    with pytest.raises(HardwareError):
        renumber_splitter(d, "710-LX500-1", 2)


def test_renumber_splitter_rejects_out_of_range():
    d = DMPDesign()
    add_splitter(d, "LX")
    with pytest.raises(HardwareError):
        renumber_splitter(d, "710-LX500-1", 0)
    with pytest.raises(HardwareError):
        renumber_splitter(d, "710-LX500-1", MAX_SPLITTERS_PER_TYPE + 1)


def test_renumber_splitter_types_are_independent():
    d = DMPDesign()
    add_splitter(d, "LX")               # 710-LX500-1
    kp = add_splitter(d, "KP")          # 710-KP-1
    renumber_splitter(d, "710-LX500-1", 2)   # KP-1 must not block LX->2
    assert {s.id for s in d.splitters} == {"710-LX500-2", "710-KP-1"}


# -------- keypads --------

def test_add_keypad_numbering_and_capacity():
    d = DMPDesign(keypads=[Keypad(number=1, source="MSP")])
    kp = add_keypad(d, location="MAIN ENTRY", source="710-KP-1", global_keypad=True)
    assert kp.number == 2 and kp.global_keypad
    while len(d.keypads) < MAX_KEYPADS:
        add_keypad(d)
    with pytest.raises(HardwareError):
        add_keypad(d)


def test_remove_keypad_scrubs_outputs():
    d = DMPDesign(keypads=[Keypad(number=3, source="MSP")])
    d.splitters.append(Splitter(id="710-KP-1", splitter_type="KP",
                                outputs=["KEYPAD #3", "Spare", "Spare"]))
    remove_keypad(d, 3)
    assert d.keypads == []
    assert d.splitters[0].outputs == ["Spare", "Spare", "Spare"]


def test_add_expander_absorbs_orphan_block_zones():
    """Real worksheets carry stray SPARE/PS zone rows beyond the installed
    expanders (DARBY has 47). Adding an expander into such a block must
    replace them — duplicates corrupt the zone grid and Master write."""
    d = _design_with_expanders(1)
    # Orphans sitting in module 2's block (Z517-532), owned by no RSP
    d.zones.append(ZoneInfo(number=517, location="SPARE", device_type="Spare", partition=1))
    d.zones.append(ZoneInfo(number=531, location="OLD STORAGE", device_type="Motion", partition=1))

    orphans = block_orphans(d, 2)
    assert {z.number for z in orphans} == {517, 531}

    rsp = add_expander(d, "714-16")
    assert rsp.number == 2
    numbers = [z.number for z in d.zones]
    assert len(numbers) == len(set(numbers)), "duplicate zone numbers"
    by_num = {z.number: z for z in d.zones}
    assert by_num[517].location == "SPARE"          # fresh row, not the orphan
    assert by_num[531].location == "PS-2: A/C LOSS"  # orphan replaced


def test_add_expander_materializes_zones_from_master():
    """Designs parsed from an xlsx with uncached Point Info formulas carry
    zone data only in master_zones; adding hardware must not let the next
    master re-sync wipe it (regression: 48 DARBY rooms vanished)."""
    from parse_dmp_worksheet import Zone
    from session import sync_master_zones
    d = DMPDesign(
        rsps=[RSP(number=1, location="FACP", zones=list(range(501, 517)))],
        power_supplies=[PowerSupply(number=1, location="FACP")],
        master_zones=[
            Zone(number=501, description="FACP ROOM", rsp_number=1),
            Zone(number=502, description="LIBRARY", rsp_number=1),
            Zone(number=515, description="PS-1: A/C LOSS", rsp_number=1, is_ps_ac=True),
            Zone(number=516, description="PS-1: BATT. TRBL", rsp_number=1, is_ps_batt=True),
        ],
        master_zones_source="master",
    )
    assert d.zones == []
    add_expander(d, "714-8")
    sync_master_zones(d)
    descs = {z.number: z.description for z in d.master_zones}
    assert descs[501] == "FACP ROOM" and descs[502] == "LIBRARY"   # preserved
    assert descs[523] == "PS-2: A/C LOSS"                          # new expander


# -------- existing_locations (autocomplete source) --------

def test_existing_locations_dedupes_case_insensitively():
    d = DMPDesign(
        rsps=[RSP(number=1, location="FACP ROOM")],
        power_supplies=[PowerSupply(number=1, location="facp room")],  # paired RSP/PS
        splitters=[Splitter(id="710-LX500-1", splitter_type="LX", location="BLDG A IDF")],
        keypads=[Keypad(number=1, source="MSP", location="MAIN ENTRY"),
                 Keypad(number=2, source="MSP", location=None)],       # blank skipped
    )
    locs = existing_locations(d)
    assert locs == ["BLDG A IDF", "FACP ROOM", "MAIN ENTRY"]  # sorted, one FACP


def test_existing_locations_empty_design():
    assert existing_locations(DMPDesign()) == []


# -------- cascade reporting (snapshot_refs / diff_refs) --------

def test_diff_refs_flags_scrubbed_outputs_on_expander_removal():
    d = _design_with_expanders(2)
    d.splitters.append(Splitter(id="710-LX500-1", splitter_type="LX",
                                outputs=["RSP-1", "RSP-2", "Spare"]))
    before = snapshot_refs(d)
    remove_expander(d, 2)
    changes = diff_refs(before, snapshot_refs(d))
    assert len(changes) == 1
    assert changes[0].tab == "SPLITTERS"
    assert "710-LX500-1 output 2" in changes[0].message and "RSP-2" in changes[0].message


def test_diff_refs_flags_blanked_keypad_source_not_removed_splitter():
    d = DMPDesign()
    s1 = add_splitter(d, "LX")
    s2 = add_splitter(d, "LX")
    s1.outputs = ["To 710-LX500-2", "RSP-1", "Spare"]
    d.keypads.append(Keypad(number=2, source="710-LX500-2", location="HALL"))
    before = snapshot_refs(d)
    remove_splitter(d, "710-LX500-2")
    changes = diff_refs(before, snapshot_refs(d))
    tabs = sorted(c.tab for c in changes)
    # s1's "To …" output → Spare (SPLITTERS) and the keypad source → blank (KEYPADS);
    # the removed splitter itself is not reported.
    assert tabs == ["KEYPADS", "SPLITTERS"]
    assert any("Keypad #2" in c.message for c in changes)


def test_diff_refs_empty_when_nothing_dangles():
    d = _design_with_expanders(1)
    before = snapshot_refs(d)
    add_expander(d, "714-16")          # adds never dangle
    assert diff_refs(before, snapshot_refs(d)) == []


# -------- round-trips --------

def test_session_round_trip_preserves_model():
    import json
    from session import design_to_dict, design_from_dict
    d = DMPDesign()
    add_expander(d, "714-8")
    restored = design_from_dict(json.loads(json.dumps(design_to_dict(d))))
    assert restored.rsps[0].model == "714-8"
    assert restored.rsps == d.rsps


@pytest.mark.skipif(not DMP_TEMPLATE.exists(),
                    reason="DMP worksheet template fixture not present")
def test_write_parse_round_trip_714_8_and_gap(tmp_path):
    from generate_dmp_ws import write_dmp_xlsx
    from parse_dmp_worksheet import parse_dmp_worksheet
    import openpyxl

    d = DMPDesign(site_info=SiteInfo(school_name="HW ROUNDTRIP"))
    add_expander(d, "714-16")
    add_expander(d, "714-8")
    add_expander(d, "714-16")
    remove_expander(d, 2)               # gap at module 2
    add_expander(d, "714-8")            # reuses module 2

    out = tmp_path / "hw.xlsx"
    write_dmp_xlsx(d, DMP_TEMPLATE, out)

    parsed = parse_dmp_worksheet(out)
    models = {r.number: r.model for r in parsed.rsps}
    assert models == {1: "714-16", 2: "714-8", 3: "714-16"}
    assert next(r for r in parsed.rsps if r.number == 2).zones == list(range(517, 525))

    # Point Info sheets kept up to max module number
    wb = openpyxl.load_workbook(out)
    info_sheets = [s for s in wb.sheetnames if "Point Info" in s]
    assert len(info_sheets) == 3


@pytest.mark.skipif(not DMP_TEMPLATE.exists(),
                    reason="DMP worksheet template fixture not present")
def test_persistent_gap_keeps_later_modules_sheets(tmp_path):
    """Modules {1,3} (module 2 removed, never refilled): sheet 3 must survive
    so module 3's hard-wired Master stride stays aligned."""
    from generate_dmp_ws import write_dmp_xlsx
    import openpyxl

    d = DMPDesign(site_info=SiteInfo(school_name="GAP TEST"))
    add_expander(d, "714-16")
    add_expander(d, "714-16")
    add_expander(d, "714-16")
    remove_expander(d, 2)

    out = tmp_path / "gap.xlsx"
    write_dmp_xlsx(d, DMP_TEMPLATE, out)
    wb = openpyxl.load_workbook(out)
    info_sheets = sorted(s for s in wb.sheetnames if "Point Info" in s)
    assert len(info_sheets) == 3, info_sheets  # sheets 1-3 kept, 4-15 trimmed
