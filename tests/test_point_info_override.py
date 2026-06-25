"""Tests for the Point Info -> Master override fold-in.

Field techs sometimes type room corrections straight into a Point Info
'LOCATION OF DEVICES' cell, overwriting its `=Master!B{row}` formula, but never
update the Master sheet. parse_dmp_worksheet treats those hand-typed literals as
authoritative (Point Info wins) and folds them into master_zones.

Run: pytest tests/test_point_info_override.py
"""
from pathlib import Path
import sys

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
SCRIPTS = REPO_ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))

from parse_dmp_worksheet import (  # noqa: E402
    RSP,
    Zone,
    _apply_overrides_to_master_zones,
    _point_info_overrides,
    parse_dmp_worksheet,
)


def _wb_with_point_info(b_override=None):
    """Minimal workbook: a Master sheet (Z501..Z516) and one Point Info sheet whose
    columns are `=Master!A{row}` / `=Master!B{row}` formulas, except b_override =
    {point_info_row: literal_text} which simulates a hand-typed override."""
    b_override = b_override or {}
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    master = wb.create_sheet("Master")
    master["A1"] = "ZONE #"
    master["B1"] = "ZONE DESCRIPTION"
    for i in range(16):
        zone = 501 + i
        mrow = 2 + i
        master.cell(mrow, 1).value = f"Z{zone}"
        master.cell(mrow, 2).value = "SPARE" if zone == 510 else f"CLASSROOM {i + 1}"

    pi = wb.create_sheet("DMP 714-16 Point Info (1)")
    pi["A1"] = "POINT INFORMATION"
    pi["A3"] = "DMP POINTS"
    pi["B3"] = "LOCATION OF DEVICES"
    for i in range(16):
        mrow = 2 + i
        pirow = 4 + i
        pi.cell(pirow, 1).value = f"=Master!A{mrow}"
        pi.cell(pirow, 2).value = b_override.get(pirow, f"=Master!B{mrow}")
    return wb


def test_detects_literal_override():
    # Z510 is Point Info row 13 (master row 11); type a room over its formula.
    wb = _wb_with_point_info(b_override={13: "PRINCIPAL OFFICE"})
    assert _point_info_overrides(wb) == {510: "PRINCIPAL OFFICE"}


def test_all_formula_yields_no_overrides():
    """A normal, untouched workbook (every col-B cell still a formula) has none."""
    assert _point_info_overrides(_wb_with_point_info()) == {}


def test_non_master_a_formula_is_ignored():
    """Legacy layouts (col A not `=Master!A{row}`) are left to the normal parse."""
    wb = _wb_with_point_info(b_override={13: "PRINCIPAL OFFICE"})
    wb["DMP 714-16 Point Info (1)"]["A13"].value = "Z510"  # literal, not a formula
    assert _point_info_overrides(wb) == {}


def test_apply_clears_spare_and_rederives_rsp():
    rsps = [RSP(number=1, location="Bldg A", zones=list(range(501, 517)))]
    master_zones = [Zone(number=510, description="SPARE", rsp_number=None, is_spare=True)]
    _apply_overrides_to_master_zones(master_zones, {510: "PRINCIPAL OFFICE"}, rsps)
    z = master_zones[0]
    assert z.description == "PRINCIPAL OFFICE"
    assert z.is_spare is False
    assert z.rsp_number == 1


def test_apply_restores_ps_flags():
    rsps = [RSP(number=2, location="Bldg B", zones=list(range(517, 533)))]
    master_zones = [Zone(number=531, description="SPARE", rsp_number=None, is_spare=True)]
    _apply_overrides_to_master_zones(master_zones, {531: "PS-2: A/C LOSS"}, rsps)
    z = master_zones[0]
    assert z.is_ps_ac and not z.is_spare and z.rsp_number == 2


def test_parse_folds_override_into_master_zones(tmp_path):
    """End to end: a workbook with a hand-typed Point Info override parses with that
    value (not the stale Master 'SPARE') in master_zones."""
    wb = _wb_with_point_info(b_override={13: "PRINCIPAL OFFICE"})
    exp = wb.create_sheet("DMP 714 Exp Mod")
    exp["A4"], exp["B4"], exp["C4"], exp["D4"] = "DMP 714-16 #", 1, "501 - 516", "Bldg A"
    path = tmp_path / "ws.xlsx"
    wb.save(path)

    design = parse_dmp_worksheet(path)
    z510 = {z.number: z for z in design.master_zones}[510]
    assert z510.description == "PRINCIPAL OFFICE"
    assert z510.is_spare is False
    assert z510.rsp_number == 1
