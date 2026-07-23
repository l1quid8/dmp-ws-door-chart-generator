"""Every chart block's title reads the school name, whatever block it is.

The door chart template stamps a three-line title (school / address / city) on
each chart block via =Header!B3, =Header!B4, =Header!B5. The template was built
by copy-pasting the block down each sheet with RELATIVE references, so blocks
below the third pointed past B5 into empty Header cells (=Header!B39, B75, ...),
printing a blank title. inject() repairs every drifted reference back to B3/B4/B5.

Run: pytest tests/test_door_chart_headers.py
"""
from pathlib import Path
import re
import sys
import zipfile

import openpyxl
import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent
SCRIPTS = REPO_ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))

from inject_door_chart import inject  # noqa: E402
from test_door_chart_consolidation import _kp, _lx, _splitter_design  # noqa: E402

DOOR_CHART_TEMPLATE = REPO_ROOT / "door_chart_template_blank.xlsx"
PRESENTATION_SHEETS = ["Terminal Cans", "RSPs", "Power Supplies", "LX-KP-710s"]

pytestmark = pytest.mark.skipif(not DOOR_CHART_TEMPLATE.exists(),
                                reason="door chart template fixture not present")

_HEADER_REF = re.compile(r"=Header!B(\d+)")


def _many_splitter_design():
    """7 LX + 3 KP — fills five LX-KP-710s blocks, exposing the drifted titles."""
    return _splitter_design(*(_lx(i) for i in range(1, 8)), *(_kp(i) for i in range(1, 4)))


def _inject(tmp_path, design):
    out = tmp_path / "door_chart.xlsx"
    inject(DOOR_CHART_TEMPLATE, design, out)
    return out


def _header_refs(ws):
    return [(c.coordinate, int(m.group(1)))
            for row in ws.iter_rows() for c in row
            if isinstance(c.value, str) and (m := _HEADER_REF.match(c.value.strip()))]


def test_template_itself_has_drifted_refs():
    """Guard the premise: the raw template really does point past B5."""
    wb = openpyxl.load_workbook(DOOR_CHART_TEMPLATE)
    drifted = [(tab, coord, n) for tab in PRESENTATION_SHEETS
               for coord, n in _header_refs(wb[tab]) if n > 5]
    assert drifted, "template no longer has drifted header refs — test is stale"


def test_all_header_refs_point_at_b3_b5(tmp_path):
    out = _inject(tmp_path, _many_splitter_design())
    wb = openpyxl.load_workbook(out)
    for tab in PRESENTATION_SHEETS:
        for coord, n in _header_refs(wb[tab]):
            assert 3 <= n <= 5, f"{tab}!{coord} still points at Header!B{n}"


def test_titles_resolve_to_the_school_name(tmp_path):
    """B3/B4/B5 carry the real site info, so every fixed ref renders it."""
    design = _many_splitter_design()
    design.site_info.school_name = "HAYNES CHARTER ES"
    out = _inject(tmp_path, design)
    wb = openpyxl.load_workbook(out)
    hdr = wb["Header"]
    assert hdr["B3"].value == "HAYNES CHARTER ES"
    # Every title cell references a populated Header row (name at minimum).
    name_refs = [(tab, coord) for tab in PRESENTATION_SHEETS
                 for coord, n in _header_refs(wb[tab]) if n == 3]
    assert name_refs, "no school-name title cells survived"


def test_good_blocks_are_untouched(tmp_path):
    """The first three blocks were already correct; the fix must not disturb them."""
    out = _inject(tmp_path, _many_splitter_design())
    zf = zipfile.ZipFile(out)
    xml = zf.read("xl/worksheets/sheet6.xml").decode()  # LX-KP-710s
    # Block 1 title still resolves to B3/B4/B5 with its original style.
    assert re.search(r'<c r="B2"[^>]*><f>Header!B3</f>', xml)
    assert re.search(r'<c r="B4"[^>]*><f>Header!B5</f>', xml)


def test_multi_sheet_full_generation_still_valid(tmp_path):
    """openpyxl re-parse validates the rewritten XML across every sheet + tables."""
    out = _inject(tmp_path, _many_splitter_design())
    wb = openpyxl.load_workbook(out)
    for tab in PRESENTATION_SHEETS:
        assert wb[tab].tables
