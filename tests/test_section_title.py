"""Splitter-topology SECTION TITLE injection (door chart Master col C).

The block header shown per splitter on the LX-KP-710s sheet must read
'<location> - 710 Splitter <bus> - <id>' (LX has a bus number from the slot id;
KP reads 'Keypad Bus'). Previously col C carried joined input descriptions.

Run: pytest tests/test_section_title.py
"""
from pathlib import Path
import sys

import openpyxl
import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent
SCRIPTS = REPO_ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))

from parse_dmp_worksheet import DMPDesign, Splitter  # noqa: E402
from inject_door_chart import format_section_title, inject  # noqa: E402

DOOR_CHART_TEMPLATE = REPO_ROOT / "door_chart_template_blank.xlsx"


def _lx(num, loc="ADMIN BUILDING", bus=500):
    return Splitter(id=f"710-LX{bus}-{num}", splitter_type="LX", location=loc)


def _kp(num, loc="ADMIN BUILDING"):
    return Splitter(id=f"710-KP-{num}", splitter_type="KP", location=loc)


# -------- unit --------

def test_lx_section_title():
    assert format_section_title(_lx(1)) == \
        "ADMIN BUILDING - 710 Splitter LX Bus 500 - 710-LX500-1"


def test_lx_bus_number_from_id():
    assert format_section_title(_lx(2, bus=600)) == \
        "ADMIN BUILDING - 710 Splitter LX Bus 600 - 710-LX600-2"


def test_kp_section_title_uses_keypad_bus():
    assert format_section_title(_kp(1)) == \
        "ADMIN BUILDING - 710 Splitter Keypad Bus - 710-KP-1"


def test_missing_location_degrades_gracefully():
    assert format_section_title(Splitter(id="710-LX500-1", splitter_type="LX", location=None)) == \
        "710 Splitter LX Bus 500 - 710-LX500-1"


# -------- integration --------

@pytest.mark.skipif(not DOOR_CHART_TEMPLATE.exists(),
                    reason="door chart template fixture not present")
def test_inject_writes_section_titles(tmp_path):
    design = DMPDesign()
    design.site_info.school_name = "TEST SCHOOL"
    design.splitters = [_lx(1, loc="MAIN BUILDING FACP ROOM"), _kp(1, loc="MAIN OFFICE")]
    out = tmp_path / "door_chart.xlsx"
    inject(DOOR_CHART_TEMPLATE, design, out)

    m = openpyxl.load_workbook(out, data_only=True)["Master"]
    # Template slot ids live in col A rows 29-63; find each splitter's row, check col C.
    by_id = {m.cell(r, 1).value: r for r in range(29, 64) if m.cell(r, 1).value}
    assert m.cell(by_id["710-LX500-1"], 3).value == \
        "MAIN BUILDING FACP ROOM - 710 Splitter LX Bus 500 - 710-LX500-1"
    assert m.cell(by_id["710-KP-1"], 3).value == \
        "MAIN OFFICE - 710 Splitter Keypad Bus - 710-KP-1"
