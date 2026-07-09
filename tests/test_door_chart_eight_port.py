"""Door chart support for 8-port (714-8) zone expanders.

The door chart's data Master sheet is a clean CONTIGUOUS zone list (zone N at
zone_to_master_row(N)). The presentation tabs (Terminal Cans, RSPs, Power Supplies) read
the Master at fixed block positions, so inject() retargets each RSP's block to its real
contiguous rows, reshapes 8-port blocks (8 data rows -> AUX POWER -> blanks), and blanks
unused module blocks. An 8-port module must NOT shift later modules off their blocks.

Run: pytest tests/test_door_chart_eight_port.py
"""
from pathlib import Path
import re
import sys
import zipfile

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent
SCRIPTS = REPO_ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))

from parse_dmp_worksheet import DMPDesign, RSP, Zone  # noqa: E402
from inject_door_chart import (  # noqa: E402
    rsp_block_anchor, zone_to_master_row, inject, _find_cell, _blank_cell,
)

DOOR_CHART_TEMPLATE = REPO_ROOT / "door_chart_template_blank.xlsx"


def _mixed_design():
    """RSP1=714-16 (501-516), RSP2=714-8 (517-524), RSP3=714-16 (525-540) — packed."""
    design = DMPDesign()
    design.site_info.school_name = "TEST SCHOOL"
    for num, zrange in [(1, range(501, 517)), (2, range(517, 525)), (3, range(525, 541))]:
        zones = list(zrange)
        design.rsps.append(RSP(number=num, location=f"BLDG {num}", zones=zones,
                               model="714-8" if len(zones) < 16 else "714-16"))
        for i, zn in enumerate(zones):
            if i == len(zones) - 2:
                z = Zone(number=zn, description=f"PS-{num}: A/C LOSS", rsp_number=num, is_ps_ac=True)
            elif i == len(zones) - 1:
                z = Zone(number=zn, description=f"PS-{num}: BATT. TRBL", rsp_number=num, is_ps_batt=True)
            elif i >= 4:
                z = Zone(number=zn, description="SPARE", rsp_number=None, is_spare=True)
            else:
                z = Zone(number=zn, description=f"ROOM {zn}", rsp_number=num)
            design.master_zones.append(z)
    return design


# -------- pure mapping unit tests --------

def test_block_anchor_bus_aware():
    assert rsp_block_anchor(1) == 67
    assert rsp_block_anchor(2) == 83
    assert rsp_block_anchor(3) == 99
    assert rsp_block_anchor(7) == 167  # bus jump (Z601)


def test_zone_to_master_row_contiguous():
    assert zone_to_master_row(501) == 67
    assert zone_to_master_row(525) == 91   # RSP3 first zone in a packed design
    assert zone_to_master_row(540) == 106


def test_find_cell_does_not_overmatch_styled_self_closing():
    """A styled self-closing cell must match only itself, so blanking it can't swallow
    the adjacent formula cell (the XML-surgery cell finder's load-bearing invariant)."""
    xml = ('<row r="5"><c r="B5" s="3"/><c r="C5" s="3"/>'
           '<c r="D5" s="4"><f>Master!D81</f><v>X</v></c></row>')
    assert _find_cell(xml, "B5").group(0) == '<c r="B5" s="3"/>'
    out = _blank_cell(xml, "B5")
    assert "Master!D81" in out and '<c r="C5" s="3"/>' in out


# -------- integration through inject() --------

def _block_first_zones(tab, master_col, data_off, zone_is_next_col):
    """For each header (=Master!{col}{n}) in reading order, resolve the Master!A row its
    first data cell points at → ('A'-row int)."""
    hdrs = sorted((c.row, c.column) for row in tab.iter_rows() for c in row
                  if isinstance(c.value, str) and re.match(r"=Master!%s\d+$" % master_col, c.value))
    out = []
    for R, col in hdrs:
        zcol = openpyxl.utils.get_column_letter(col + (1 if zone_is_next_col else 0))
        v = tab[f"{zcol}{R + data_off}"].value
        mm = re.search(r"Master!A(\d+)", v) if isinstance(v, str) else None
        out.append(int(mm.group(1)) if mm else None)
    return out


@pytest.mark.skipif(not DOOR_CHART_TEMPLATE.exists(),
                    reason="door chart template fixture not present")
def test_inject_contiguous_master_and_mapping(tmp_path):
    out = tmp_path / "door_chart.xlsx"
    inject(DOOR_CHART_TEMPLATE, _mixed_design(), out)
    wb = openpyxl.load_workbook(out, data_only=False)
    m = wb["Master"]

    # Master is a clean contiguous list, no gaps, no junk tail.
    assert m["A83"].value == "Z517"
    assert m["A91"].value == "Z525"        # RSP3 immediately follows RSP2 (no 8-port gap)
    assert m["A106"].value == "Z540"
    assert m["A107"].value is None and m["B107"].value is None  # nothing past the real zones

    # Terminal Cans + RSPs: each block points to its RSP's real contiguous first row.
    # RSP1→67 (Z501), RSP2→83 (Z517), RSP3→91 (Z525) — RSP3 did NOT get pushed to 99.
    assert _block_first_zones(wb["Terminal Cans"], "D", 2, False)[:3] == [67, 83, 91]
    assert _block_first_zones(wb["RSPs"], "C", 3, True)[:3] == [67, 83, 91]

    # Only the 3 real blocks remain populated; unused module blocks are blanked.
    tc_hdrs = [c.value for row in wb["Terminal Cans"].iter_rows() for c in row
               if isinstance(c.value, str) and re.match(r"=Master!D\d+$", c.value)]
    assert len(tc_hdrs) == 3


@pytest.mark.skipif(not DOOR_CHART_TEMPLATE.exists(),
                    reason="door chart template fixture not present")
def test_inject_keeps_excel_table_headers(tmp_path):
    """The presentation sheets carry Excel Tables; emptying a table's header row (e.g. while
    blanking an unused block) makes Excel flag the file for repair. Every table's header-row
    cells must stay populated after inject()."""
    out = tmp_path / "door_chart.xlsx"
    inject(DOOR_CHART_TEMPLATE, _mixed_design(), out)
    zt = zipfile.ZipFile(DOOR_CHART_TEMPLATE)
    zf = zipfile.ZipFile(out)

    def table_header_spans(sheet_num):
        # Read the OUTPUT's rels: consolidation drops the tables of truncated blocks,
        # so only surviving tables still name header cells the sheet must keep.
        rels = zf.read(f"xl/worksheets/_rels/sheet{sheet_num}.xml.rels").decode()
        spans = []
        for t in re.findall(r"tables/(table\d+\.xml)", rels):
            m = re.search(r'ref="([A-Z]+)(\d+):([A-Z]+)\d+"', zf.read("xl/tables/" + t).decode())
            if m:
                spans.append((int(m.group(2)), m.group(1), m.group(3)))
        return spans

    for sheet_num in (3, 4, 5, 6):  # Terminal Cans, RSPs, Power Supplies, LX-KP-710s
        out_xml = zf.read(f"xl/worksheets/sheet{sheet_num}.xml").decode()
        tpl_xml = zt.read(f"xl/worksheets/sheet{sheet_num}.xml").decode()
        for header_row, c1, c2 in table_header_spans(sheet_num):
            for col in range(column_index_from_string(c1), column_index_from_string(c2) + 1):
                ref = f"{get_column_letter(col)}{header_row}"
                tpl = re.search(r'<c r="%s"[^>]*?(?:/>|>.*?</c>)' % ref, tpl_xml, re.S)
                if not (tpl and ("<v>" in tpl.group(0) or 't="s"' in tpl.group(0))):
                    continue  # template cell wasn't a populated header; skip
                got = re.search(r'<c r="%s"[^>]*?(?:/>|>.*?</c>)' % ref, out_xml, re.S)
                assert got and ("<v>" in got.group(0) or 't="s"' in got.group(0)), \
                    f"table header cell {ref} on sheet{sheet_num} was blanked"


@pytest.mark.skipif(not DOOR_CHART_TEMPLATE.exists(),
                    reason="door chart template fixture not present")
def test_inject_eight_port_reshape_and_power_supplies(tmp_path):
    out = tmp_path / "door_chart.xlsx"
    inject(DOOR_CHART_TEMPLATE, _mixed_design(), out)
    wb = openpyxl.load_workbook(out, data_only=False)
    tc = wb["Terminal Cans"]

    # RSP2 (8-port) is the right block of group 1 (header F7): 8 data rows then AUX POWER.
    assert str(tc["F16"].value).startswith("=Master!A")   # 8th (last) real data row
    assert tc["F17"].value == "AUX POWER"                  # AUX lifted up
    assert tc["F18"].value is None                         # blanks below

    # Power Supplies reads RSP2's real supervisory rows (Z523/Z524 → Master rows 89/90),
    # not the 16-port slot positions.
    ps = wb["Power Supplies"]
    refs = {c.value for row in ps.iter_rows() for c in row if isinstance(c.value, str)}
    assert any("Master!A89" in v for v in refs)
    assert any("Master!A90" in v for v in refs)
